from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
import time
import sys
import openai


class DataFile:
    wb_data_grab = load_workbook("AccreditationData.xlsx")
    wb_nces_grab = load_workbook("Data_3-14-2023---623.xlsx")
    ws_data_grab = wb_data_grab["InstituteCampuses"]
    ws_nces_grab = wb_nces_grab["Data_3-14-2023---623"]

    def __init__(self, raw_file, sheet_name, abbrev):
        self.raw_file = raw_file
        self.sheet_name = sheet_name
        self.abbrev = abbrev
        self.wb_uasys = load_workbook(raw_file)
        self.ws_uasys = self.wb_uasys[sheet_name]

    @staticmethod
    def raw_file_check(raw_file):
        wrong_input = raw_file.find(".xlsx")
        if wrong_input == -1:
            print("Be sure to add .xlsx to end of file location!")
            raw_file = input("File location is explorer(.xlsx)")
            return raw_file

    @staticmethod
    def has_numbers(input_string: str):
        return any(char.isdigit() for char in input_string)

    @classmethod
    def reconcile_institution(cls, wb_uasys, ws_uasys, raw_file, ws_data_grab, ws_nces_grab):
        # Inputs Autogen in field cells
        for cell in ws_uasys['Q']:
            try:
                if cell.value is None:
                    ws_uasys['Q' + str(cell.row)].value = "AutoGen"
            except AttributeError:
                print(cell + ' is read only!')
            except TypeError:
                print('Cell is read only!')
            except:
                print('Unknown error')
        # Get inst_po_box_line for primary_institution_name from LocationName -> address
        for cell in ws_uasys['U']:
            primary_institution_name = str(cell.value).upper()
            cell_prev = int(cell.row) - 1
            try:
                if cell_prev != 0 and primary_institution_name != ws_uasys['U' + str(cell_prev)].value.upper():
                    print("----------------------------------")
                    print(primary_institution_name)
                    for grab in ws_data_grab['D']:
                        location_name = str(grab.value)
                        if location_name.upper() == primary_institution_name:
                            address = str(ws_data_grab['H' + str(grab.row)].value)
                            if address.find('P.O'):
                                found = re.search("x(.+?),", address)
                                if not found:
                                    continue
                                else:
                                    number_pobox = found.group(1)
                                    inst_po_box_line = str('PO Box' + str(number_pobox))
                                    print('Found: ' + inst_po_box_line)
                                    ws_uasys['X' + str(cell.row)].value = inst_po_box_line
            except AttributeError:
                print("----------------------------------")
                print('NoneType for: ' + str(cell.value))
            except TypeError:
                print('NoneType')
            except:
                print('Unknown error')
        # If INST_COUNTRY_CODE is blank then assign USA
        for cell in ws_uasys['AA']:
            try:
                if cell.value is None:
                    ws_uasys['AA' + str(cell.row)].value = "USA"
            except AttributeError:
                print('cell is read only!')
            except TypeError:
                print('Cell is read only!')
            except:
                print('Unknown error')
        # Move/delete substrings from inst_address_line_1 and moving them into respective column row
        for cell in ws_uasys['V']:
            governing_address = str(cell.value).split()
            for index in range(len(governing_address)):
                word = governing_address[index]
                if word == 'Ste' or word == 'Ste.' or word == 'Unit' or word == 'PO' or word == 'Suite' \
                        or word == 'Building':
                    inst_address_line_1 = str(' '.join(governing_address[index:len(governing_address)]))
                    found_pobox = inst_address_line_1.find('PO Box')
                    if found_pobox == -1:
                        ws_uasys['W' + str(cell.row)].value = inst_address_line_1.upper()
                    else:
                        ws_uasys['X' + str(cell.row)].value = inst_address_line_1
                        ws_uasys['W' + str(cell.row)].value = 'N/A'
                    address_line_1 = str(cell.value)
                    phrase_removal = address_line_1.find(inst_address_line_1)
                    if phrase_removal != -1:
                        ws_uasys['V' + str(cell.row)].value = address_line_1.strip(inst_address_line_1)
                elif word == 'Floor' or word == 'Fl':
                    floor_num = index - 1
                    inst_address_line_1 = str(' '.join(governing_address[floor_num:len(governing_address)]))
                    ws_uasys['W' + str(cell.row)].value = inst_address_line_1.upper()
                    address_line_1 = str(cell.value)
                    phrase_removal = address_line_1.find(inst_address_line_1)
                    if phrase_removal != -1:
                        ws_uasys['V' + str(cell.row)].value = address_line_1.strip(inst_address_line_1)
        # If INST_ADDRESS_LINE_2 is blank then assign the cell N/A
        for cell in ws_uasys['W']:
            try:
                if cell.value is None:
                    ws_uasys['W' + str(cell.row)].value = "N/A"
            except AttributeError:
                print('Cell is read only!')
            except TypeError:
                print('Cell is read only!')
            except:
                print('Unknown error')
        # Check to see if institution is inactive/closed according to NCES database
        for cell in ws_uasys['U']:
            organization_name = str(cell.value)
            cell_prev = int(cell.row) - 1
            try:
                if cell_prev != 0 and organization_name != ws_uasys['U' + str(cell_prev)].value.upper():
                    for look in ws_nces_grab['B']:
                        nces_institution = str(look.value)
                        if nces_institution.upper() == organization_name.upper():
                            institution_closed = ws_nces_grab['W' + str(look.row)].value
                            found_two = str(institution_closed).find('-2')
                            if found_two < 0:
                                ws_uasys['AI' + str(cell.row)].value = institution_closed
                                ws_uasys['AH' + str(cell.row)].value = 'Y'
            except AttributeError:
                print("----------------------------------")
                print('NoneType for: ' + str(cell.value))
            except TypeError:
                print('NoneType')
            except:
                print('Unknown error')
        # if INST_RECORD_SOURCE is blank then assign N/A
        for cell in ws_uasys['AJ']:
            try:
                if cell.value is None:
                    ws_uasys['AJ' + str(cell.row)].value = "N/A"
            except AttributeError:
                print('cell is read only!')
            except TypeError:
                print('Cell is read only!')
            except:
                print('Unknown error')
        print('Done!')
        wb_uasys.save(raw_file)

    @classmethod
    def clean_institution(cls, wb_uasys, ws_uasys, raw_file):
        for cell in ws_uasys['R']:
            try:
                if cell.value is None:
                    ws_uasys['R' + str(cell.row)].value = "NULL"
            except:
                print('Error with cell')
        for cell in ws_uasys['S']:
            try:
                if cell.value is None:
                    ws_uasys['S' + str(cell.row)].value = "NULL"
            except:
                print('Error with cell')
        for cell in ws_uasys['T']:
            try:
                if cell.value is None:
                    ws_uasys['T' + str(cell.row)].value = "NULL"
            except:
                print('Error with cell')
        yellow = 'FFFF00'
        red = 'FF6666'
        y_highlight = PatternFill(patternType='solid', fgColor=yellow)
        r_highlight = PatternFill(patternType='solid', fgColor=red)
        for cell in ws_uasys['Q']:
            try:
                if cell.value is None:
                    ws_uasys['Q' + str(cell.row)].value = 'AutoGen'
            except:
                print('Error with cell')
        for cell in ws_uasys['R']:
            try:
                if cell.row >= 3:
                    gov_id = str(cell.value)
                    gov_id_oped = str(ws_uasys['S' + str(cell.row)].value)
                    gov_id_iped = str(ws_uasys['T' + str(cell.row)].value)
                    if not gov_id.isnumeric():
                        ws_uasys['R' + str(cell.row)].fill = r_highlight
                    if not gov_id_oped.isnumeric():
                        ws_uasys['S' + str(cell.row)].fill = r_highlight
                    if not gov_id_iped.isnumeric():
                        ws_uasys['T' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['U']:
            try:
                ws_uasys['U' + str(cell.row)].value = str(cell.value).upper()
                if cell.value is None:
                    ws_uasys['U' + str(cell.row)].fill = y_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['V']:
            try:
                ws_uasys['V' + str(cell.row)].value = str(cell.value).upper()
                if cell.value is None:
                    ws_uasys['V' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['W']:
            try:
                address_one = cell.value
                if address_one is None:
                    ws_uasys['W' + str(cell.row)].value = 'N/A'
                if address_one.find('PO') != -1:
                    ws_uasys['W' + str(cell.row)].fill = y_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['X']:
            try:
                address_two = str(cell.value)
                if cell.value is None:
                    ws_uasys['X' + str(cell.row)].value = 'N/A'
                    wb_uasys.save(raw_file)
                if address_two.find('PO') == -1 and address_two != 'N/A':
                    ws_uasys['X' + str(cell.row)].fill = y_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['Y']:
            try:
                ws_uasys['Y' + str(cell.row)].value = str(cell.value).upper()
                if cell.value is None:
                    ws_uasys['Y' + str(cell.row)].fill = r_highlight
            except: print('Error with cell')
        for cell in ws_uasys['Z']:
            try:
                if cell.row >= 3:
                    region = str(cell.value)
                    if len(region) != 2:
                        ws_uasys['Z' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['AA']:
            try:
                if cell.value is None:
                    ws_uasys['AA' + str(cell.row)].value = 'USA'
            except:
                print('Error with cell')
        for cell in ws_uasys['AB']:
            try:
                if cell.value is None:
                    ws_uasys['AB' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['AC']:
            try:
                if cell.value is None:
                    ws_uasys['AC' + str(cell.row)].fill = y_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['AF']:
            if cell.row >= 3:
                try:
                    if cell.value == 'Manually Find' or cell.value is None:
                        ws_uasys['AF' + str(cell.row)].fill = r_highlight
                except:
                    print('Error with cell')
        for cell in ws_uasys['AG']:
            if cell.row >= 3:
                try:
                    if cell.value == 'Manually Find' or cell.value is None:
                        ws_uasys['AG' + str(cell.row)].fill = r_highlight
                except:
                    print('Error with cell')
        for cell in ws_uasys['AJ']:
            try:
                if cell.value is None:
                    ws_uasys['AJ' + str(cell.row)].value = "N/A"
            except:
                print('Error with cell')
        wb_uasys.save(raw_file)
        print('Done!')

    @classmethod
    def ai_institution(cls, wb_uasys, ws_uasys, raw_file):
        max_row = ws_uasys.max_row
        for cell in ws_uasys['U']:
            progress = cell.row/max_row
            sys.stdout.write('\r')
            sys.stdout.write("[%-20s] %d%%" % ('=' * int(max_row * progress), float(cell.row/max_row)*100))
            try:
                if cell.row >= 3:
                    cell_prev = int(cell.row) - 1
                    institution_name = str(cell.value)
                    municipality = str(ws_uasys['Y' + str(cell.row)].value)
                    state = str(ws_uasys['Z' + str(cell.row)].value)
                    if institution_name != ws_uasys['U' + str(cell_prev)].value and ws_uasys[
                        'AF' + str(cell.row)].value is None:
                        API_KEY = open(r"C:\Users\Wayne\Work Stuff\Data Conversion\API Key.txt").read()
                        openai.api_key = API_KEY
                        response = openai.ChatCompletion.create(
                            model="gpt-3.5-turbo",
                            messages=[
                                {"role": "system", "content": "You are a data analyst reconciling missing data."},
                                {"role": "user",
                                 "content": "Don't include the question in your response, what is the date when"
                                            "Texas State University at San Marcos, TX founded?"},
                                {"role": "assistant", "content": "1899-01-01"},
                                {"role": "user",
                                 "content": "Don't include the question in your response, what is the date when"
                                            "SAINT MARY'S COLLEGE OF CALIFORNIA at MORAGA, CA founded?"},
                                {"role": "assistant", "content": "1863-01-01"},
                                {"role": "user", "content": "If you can not find the date please respond with N/A."},
                                {"role": "assistant", "content": "N/A"},
                                {"role": "user",
                                 "content": "Don't include the question in your response, what is the date when "
                                            + institution_name + ' at ' + municipality + ', ' + state + " founded?"}
                            ]
                        )
                        reply_content = response.choices[0].message.content
                        if DataFile.has_numbers(reply_content):
                            ws_uasys['AF' + str(cell.row)].value = str(reply_content)
                        else:
                            ws_uasys['AF' + str(cell.row)].value = 'NULL'
                        wb_uasys.save(raw_file)
                        time.sleep(1)
                wb_uasys.save(raw_file)
            except openai.error.APIError:
                print('Bad Gateway')
            except openai.error.ServiceUnavailableError:
                print('Server Overload')
            except TimeoutError:
                print('Read Operation Timeout')
            except:
                print('Server overload')
            sys.stdout.flush()

        for cell in ws_uasys['U']:
            progress = cell.row/max_row
            sys.stdout.write('\r')
            sys.stdout.write("[%-20s] %d%%" % ('=' * int(max_row * progress), float(cell.row/max_row)*100))
            try:
                if cell.row >= 3:
                    cell_prev = int(cell.row) - 1
                    institution_name = str(cell.value)
                    municipality = str(ws_uasys['Y' + str(cell.row)].value)
                    state = str(ws_uasys['Z' + str(cell.row)].value)
                    if institution_name != ws_uasys['U' + str(cell_prev)].value and ws_uasys[
                        'AG' + str(cell.row)].value is None:
                        API_KEY = open(r"C:\Users\Wayne\Work Stuff\Data Conversion\API Key.txt").read()
                        openai.api_key = API_KEY
                        response = openai.ChatCompletion.create(
                            model="gpt-3.5-turbo",
                            messages=[
                                {"role": "system", "content": "You are a data analyst reconciling missing data."},
                                {"role": "user",
                                 "content": "Don't include the question in your response, When was this "
                                            "institution named Texas State University in San Marcos, TX?"},
                                {"role": "assistant", "content": "2013-01-01"},
                                {"role": "user",
                                 "content": "Don't include the question in your response, When was this "
                                            "institution named SAINT MARY'S COLLEGE OF CALIFORNIA in MORAGA, CA?"},
                                {"role": "assistant", "content": "1863-01-01"},
                                {"role": "user", "content": "If you can not find the date please respond with N/A."},
                                {"role": "assistant", "content": "N/A"},
                                {"role": "user",
                                 "content": "Don't include the question in your response, When was this "
                                            "institution named " + institution_name + ' in ' + municipality + ', '
                                            + state + "?"}
                            ]
                        )
                        reply_content = response.choices[0].message.content
                        if DataFile.has_numbers(reply_content):
                            ws_uasys['AG' + str(cell.row)].value = str(reply_content)
                        else:
                            ws_uasys['AG' + str(cell.row)].value = 'NULL'
                        wb_uasys.save(raw_file)
                        time.sleep(1)
                wb_uasys.save(raw_file)
            except openai.error.APIError:
                print('Bad Gateway')
            except openai.error.ServiceUnavailableError:
                print('Server Overload')
            except TimeoutError:
                print('Read Operation Timeout')
            except:
                print('Server Overload')
            sys.stdout.flush()

    @classmethod
    def reconcile_governing(cls, wb_uasys, ws_uasys, raw_file, abbrev, ws_data_grab, ws_nces_grab):
        # If GOVERNING_ORGANIZATION_ID is blank then assign the cell AutoGen
        for cell in ws_uasys['A']:
            try:
                if cell.value is None:
                    ws_uasys['A' + str(cell.row)].value = "AutoGen"
            except AttributeError:
                print('Cell is read only!')
            except TypeError:
                print('Cell is read only!')
            except:
                print('Unknown error')
        # Get primary institution name and compare it against cells in additional sites location name,
        # if match: access Parent
        # Name Cell and return cell data to populate Governing Org name of same row as primary institution name
        for cell in ws_uasys['U']:
            institute_name = str(cell.value)
            print("----------------------------------")
            print(institute_name)
            for grab in ws_data_grab['D']:
                location_name = str(grab.value)
                if location_name.upper() == institute_name.upper():
                    parent_name = str(ws_data_grab['E' + str(grab.row)].value)
                    if parent_name == "-" and ws_uasys['F' + str(cell.row)].value is not None:
                        ws_uasys['E' + str(cell.row)].value = institute_name
                    else:
                        ws_uasys['E' + str(cell.row)].value = parent_name
                    print('Placed in governing: ' + str(ws_uasys['E' + str(cell.row)].value))
                    print("----------------------------------")
        print("Populating associated fields.....hold on.....")
        # Get Governing_Organization_Name's DAPIP, OPE, and IPEDSID IDs from data_grab
        for cell in ws_uasys['E']:
            institution_govern = str(cell.value)
            for grab in ws_data_grab['E']:
                location_name = str(grab.value)
                if location_name.upper() == institution_govern.upper():
                    GOV_DAPID = str(ws_data_grab['F' + str(grab.row)].value)
                    ws_uasys['B' + str(cell.row)].value = GOV_DAPID
                    ws_uasys['C' + str(cell.row)].value = 'SEARCH'
                    ws_uasys['D' + str(cell.row)].value = 'SEARCH'
            dapid_check = str(ws_uasys['B' + str(cell.row)].value)
            if dapid_check == '-':
                govern_zipcode = str(ws_uasys['L' + str(cell.row)].value)
                for look in ws_nces_grab['B']:
                    match_institution = str(look.value)
                    nces_zipcode = str(ws_nces_grab['K' + str(look.row)].value)
                    if match_institution.upper() == institution_govern.upper() and govern_zipcode == nces_zipcode:
                        found_zipcode = str(ws_nces_grab['A' + str(look.row)].value)
                        ws_uasys['B' + str(cell.row)].value = found_zipcode
        # Get GOV address line 1, GOV_MUNICIPALITY, GOV postal code
        for cell in ws_uasys['E']:
            institution_govern = str(cell.value)
            for grab in ws_data_grab['D']:
                location_name = str(grab.value)
                if location_name.upper() == institution_govern.upper():
                    address_grab = str(ws_data_grab['H' + str(grab.row)].value)
                    address_grab.split(', ')
                    try:
                        if len(address_grab.split(', ')) == 1:
                            address_grab = address_grab + ", N/A, N/A, N/A, N/A, N/A, N/A"
                        if len(address_grab.split(', ')) == 2:
                            address_grab = address_grab + ", N/A, N/A, N/A, N/A, N/A"
                        if len(address_grab.split(', ')) == 2:
                            address_grab = address_grab + ", N/A, N/A, N/A, N/A, N/A"
                        if len(address_grab.split(', ')) == 3:
                            address_grab = address_grab + ", N/A, N/A, N/A, N/A"
                        if len(address_grab.split(', ')) == 4:
                            address_grab = address_grab + ", N/A, N/A, N/A"

                        GOV_ADDRESS_LINE_1, temp_LINE_2, temp_POBOX, temp_MUNI, temp_PCODE, temp1_Unknown, temp2_Unknown = address_grab.split(
                            ', ')

                        if GOV_ADDRESS_LINE_1.startswith('P.O. Box'):
                            temp_PCODE = temp_POBOX
                            temp_POBOX = GOV_ADDRESS_LINE_1
                            GOV_ADDRESS_LINE_1 = 'N/A'

                        if temp_POBOX.startswith('K'):
                            temp_POBOX = 'N/A'
                            temp_MUNI = temp_PCODE
                            temp_PCODE = temp1_Unknown

                        if temp_PCODE.startswith('P.O BOX'):
                            temp_POBOX = temp_PCODE
                            temp_PCODE = 'NULL'

                        if not temp_LINE_2.startswith('Suite'):
                            temp_MUNI = temp_LINE_2
                            temp_LINE_2 = 'N/A'
                            temp_PCODE = temp_POBOX
                            temp_POBOX = 'N/A'
                        if temp_MUNI.startswith(abbrev.upper()):
                            temp_PCODE = temp_MUNI
                            temp_MUNI = temp_POBOX
                            temp_POBOX = 'N/A'

                        GOV_ADDRESS_LINE_2 = temp_LINE_2.upper()
                        GOV_PO_BOX_LINE = temp_POBOX.strip('.')
                        GOV_MUNICIPALITY = temp_MUNI.upper()
                        GOV_POSTAL_CODE = temp_PCODE.strip(abbrev.upper())

                        ws_uasys['F' + str(cell.row)].value = GOV_ADDRESS_LINE_1
                        ws_uasys['G' + str(cell.row)].value = GOV_ADDRESS_LINE_2
                        ws_uasys['H' + str(cell.row)].value = GOV_PO_BOX_LINE
                        ws_uasys['I' + str(cell.row)].value = GOV_MUNICIPALITY
                        ws_uasys['L' + str(cell.row)].value = GOV_POSTAL_CODE
                    except ValueError:
                        ws_uasys['F' + str(cell.row)].value = 'NULL'
                        ws_uasys['G' + str(cell.row)].value = 'NULL'
                        ws_uasys['H' + str(cell.row)].value = 'NULL'
                        ws_uasys['I' + str(cell.row)].value = 'NULL'
                        ws_uasys['L' + str(cell.row)].value = 'NULL'
                    except:
                        ws_uasys['F' + str(cell.row)].value = 'NULL'
                        ws_uasys['G' + str(cell.row)].value = 'NULL'
                        ws_uasys['H' + str(cell.row)].value = 'NULL'
                        ws_uasys['I' + str(cell.row)].value = 'NULL'
                        ws_uasys['L' + str(cell.row)].value = 'NULL'
        # If GOV_STATE_REGION_SHORT is blank then assign worksheet state
        for cell in ws_uasys['J']:
            try:
                if cell.value is None:
                    ws_uasys['J' + str(cell.row)].value = abbrev.upper()
            except AttributeError:
                print('Cell is read only!')
            except TypeError:
                print('Cell is read only!')
            except:
                print('Unknown error')
        # If GOV_COUNTRY_CODE is blank then assign USA
        for cell in ws_uasys['K']:
            try:
                if cell.value is None:
                    ws_uasys['K' + str(cell.row)].value = "USA"
            except AttributeError:
                print('Cell is read only!')
            except TypeError:
                print('Cell is read only!')
            except:
                print('Unknown error')
        # Get GOV_PhoneNumberFull
        for cell in ws_uasys['E']:
            institution_govern = str(cell.value)

            for grab in ws_data_grab['D']:
                location_name = str(grab.value)
                if location_name.upper() == institution_govern.upper():
                    phoneNumber_grab = str(ws_data_grab['I' + str(grab.row)].value)
                    ws_uasys['M' + str(cell.row)].value = phoneNumber_grab

                    if ws_uasys['M' + str(cell.row)].value is None:
                        print('No phone number from Accreditation Database : Searching')
                        for look in ws_nces_grab['B']:
                            nces_institution = str(look.value)
                            if nces_institution.upper() == institution_govern.upper():
                                phoneNumber_grab = str(ws_nces_grab['L' + str(grab.row)].value)
                                ws_uasys['M' + str(cell.row)].value = phoneNumber_grab
                                print('Found phone number in NCES Database')
        # Check to see if GOV_ORG is inactive/closed according to NCES database
        for cell in ws_uasys['E']:
            institution_govern = str(cell.value)
            for look in ws_nces_grab['B']:
                nces_institution = str(look.value)
                if nces_institution.upper() == institution_govern.upper():
                    institution_closed = ws_nces_grab['W' + str(look.row)].value
                    found_two = str(institution_closed).find('-2')
                    if found_two < 0:
                        ws_uasys['O' + str(cell.row)].value = institution_closed
        # If GOV_RECORD_SOURCE is blank then assign the cell N/A
        for cell in ws_uasys['P']:
            try:
                if cell.value is None:
                    ws_uasys['P' + str(cell.row)].value = "N/A"
            except AttributeError:
                print('Cell is read only!')
            except TypeError:
                print('Cell is read only!')
            except:
                print('Unknown error')
        # if not in data_grab then search nces_grab database
        for cell in ws_uasys['E']:
            try:
                if cell.value is None:
                    search_institution = ws_uasys['U' + str(cell.row)].value
                    for look in ws_nces_grab['B']:
                        nces_institution = str(look.value)
                        if nces_institution.upper() == search_institution.upper():
                            GOV_DAPID = ws_uasys['R' + str(cell.row)].value
                            GOV_OPEID = ws_uasys['S' + str(cell.row)].value
                            GOV_IPEDID = ws_uasys['T' + str(cell.row)].value
                            PRIMARY_INSTITUTION_NAME = nces_institution.upper()
                            GOV_ADDRESS_LINE_1 = ws_nces_grab['I' + str(look.row)].value
                            GOV_MUNICIPALITY = ws_nces_grab['J' + str(look.row)].value
                            GOV_STATE_REGION_SHORT = ws_nces_grab['C' + str(look.row)].value
                            GOV_POSTAL_CODE = ws_nces_grab['K' + str(look.row)].value
                            GOV_PhoneNumberFull = ws_nces_grab['L' + str(look.row)].value

                            ws_uasys['B' + str(cell.row)].value = GOV_DAPID
                            ws_uasys['C' + str(cell.row)].value = GOV_OPEID
                            ws_uasys['D' + str(cell.row)].value = GOV_IPEDID
                            ws_uasys['E' + str(cell.row)].value = PRIMARY_INSTITUTION_NAME.upper()
                            ws_uasys['F' + str(cell.row)].value = GOV_ADDRESS_LINE_1
                            ws_uasys['I' + str(cell.row)].value = GOV_MUNICIPALITY
                            ws_uasys['J' + str(cell.row)].value = GOV_STATE_REGION_SHORT
                            ws_uasys['L' + str(cell.row)].value = GOV_POSTAL_CODE
                            ws_uasys['M' + str(cell.row)].value = GOV_PhoneNumberFull
            except AttributeError:
                print('Cell is read only!')
            except TypeError:
                print('Cell is read only!')
            except:
                print('Unknown error')
        # Move/delete substrings from GOV_ADDRESS_LINE_1 and moving them into respective column row
        for cell in ws_uasys['F']:
            governing_address = str(cell.value).split()
            for index in range(len(governing_address)):
                word = governing_address[index]
                if word == 'Ste' or word == 'Ste.' or word == 'Unit' or word == 'PO' or word == 'Suite':
                    GOV_ADDRESS_LINE_2 = str(' '.join(governing_address[index:len(governing_address)]))
                    found_pobox = GOV_ADDRESS_LINE_2.find('PO Box')
                    if found_pobox == -1:
                        ws_uasys['G' + str(cell.row)].value = GOV_ADDRESS_LINE_2.upper()
                    else:
                        ws_uasys['H' + str(cell.row)].value = GOV_ADDRESS_LINE_2
                        ws_uasys['G' + str(cell.row)].value = 'N/A'
                    ADDRESS_LINE_1 = str(cell.value)
                    phrase_removal = ADDRESS_LINE_1.find(GOV_ADDRESS_LINE_2)
                    if phrase_removal != -1:
                        ws_uasys['F' + str(cell.row)].value = ADDRESS_LINE_1.strip(GOV_ADDRESS_LINE_2)
                elif word == 'Floor' or word == 'Fl':
                    floor_num = index - 1
                    GOV_ADDRESS_LINE_2 = str(' '.join(governing_address[floor_num:len(governing_address)]))
                    ws_uasys['G' + str(cell.row)].value = GOV_ADDRESS_LINE_2.upper()
                    ADDRESS_LINE_1 = str(cell.value)
                    phrase_removal = ADDRESS_LINE_1.find(GOV_ADDRESS_LINE_2)
                    if phrase_removal != -1:
                        ws_uasys['F' + str(cell.row)].value = ADDRESS_LINE_1.strip(GOV_ADDRESS_LINE_2)
        # Move/delete substrings from GOV_POSTAL_CODE to GOV_MUNICIPALITY, GOV_MUNICIPALITY moves to GOV_ADDRESS_LINE_2
        for cell in ws_uasys['L']:
            postal_code = str(cell.value).split()
            for index in range(len(postal_code)):
                word = postal_code[index]
                if not word.isalpha():
                    if word == "N/":
                        gov_municipality = ws_uasys['I' + str(cell.row)].value.split()
                        ws_uasys['J' + str(cell.row)].value = str(gov_municipality[0])
                        ws_uasys['L' + str(cell.row)].value = str(gov_municipality[1])
                        GOV_MUNICIPALITY = ws_uasys['H' + str(cell.row)].value
                        ws_uasys['I' + str(cell.row)].value = GOV_MUNICIPALITY
                        ws_uasys['H' + str(cell.row)].value = 'N/A'
                else:
                    GOV_POSTAL_CODE = str(' '.join(postal_code[index:len(postal_code)]))
                    if GOV_POSTAL_CODE.isalpha():
                        try:
                            GOV_ADDRESS_LINE_1 = ws_uasys['I' + str(cell.row)].value
                            GOV_MUNICIPALITY = ws_uasys['L' + str(cell.row)].value
                            ws_uasys['I' + str(cell.row)].value = GOV_MUNICIPALITY
                            ws_uasys['G' + str(cell.row)].value = GOV_ADDRESS_LINE_1
                            ws_uasys['L' + str(cell.row)].value = ''
                        except AttributeError:
                            print('MergedCell object attribute value is read-only')
                    else:
                        GOV_STATE_REGION_SHORT = str(postal_code[index])
                        ws_uasys['J' + str(cell.row)].value = GOV_STATE_REGION_SHORT
                        ws_uasys['L' + str(cell.row)].value = GOV_POSTAL_CODE.strip(GOV_STATE_REGION_SHORT)
        # Check to see if institution is inactive/closed according to NCES database
        for cell in ws_uasys['E']:
            institution_govern = str(cell.value)
            cell_prev = int(cell.row) - 1
            try:
                if cell_prev != 0 and institution_govern.upper() != ws_uasys['E' + str(cell_prev)].value.upper():
                    for look in ws_nces_grab['B']:
                        nces_institution = str(look.value)
                        if nces_institution.upper() == institution_govern.upper():
                            institution_closed = ws_nces_grab['W' + str(look.row)].value
                            found_two = str(institution_closed).find('-2')
                            if found_two < 0:
                                ws_uasys['AI' + str(cell.row)].value = institution_closed
            except AttributeError:
                print("----------------------------------")
                print('NoneType for: ' + str(cell.value))
            except TypeError:
                print('NoneType')
            except:
                print('Unknown error')
        print('Done!')
        wb_uasys.save(raw_file)
        
    @classmethod
    def clean_governing(cls, wb_uasys, ws_uasys, raw_file):
        for cell in ws_uasys['B']:
            try:
                if cell.value is None:
                    ws_uasys['B' + str(cell.row)].value = "NULL"
            except:
                print('Error with cell')
        for cell in ws_uasys['C']:
            try:
                if cell.value is None:
                    ws_uasys['C' + str(cell.row)].value = "NULL"
            except:
                print('Error with cell')
        for cell in ws_uasys['D']:
            try:
                if cell.value is None:
                    ws_uasys['D' + str(cell.row)].value = "NULL"
            except:
                print('Error with cell')
        yellow = 'FFFF00'
        red = 'FF6666'
        y_highlight = PatternFill(patternType='solid', fgColor=yellow)
        r_highlight = PatternFill(patternType='solid', fgColor=red)
        for cell in ws_uasys['A']:
            if cell.row >= 3:
                try:
                    org_id = str(cell.value)
                    if org_id != 'AutoGen':
                        ws_uasys['A' + str(cell.row)].value = 'AutoGen'
                except:
                    print('Error with cell')
        for cell in ws_uasys['B']:
            try:
                if cell.row >= 3:
                    gov_id = str(cell.value)
                    gov_id_oped = str(ws_uasys['C' + str(cell.row)].value)
                    gov_id_iped = str(ws_uasys['D' + str(cell.row)].value)
                    if not gov_id.isnumeric():
                        ws_uasys['B' + str(cell.row)].fill = r_highlight
                    if not gov_id_oped.isnumeric():
                        ws_uasys['C' + str(cell.row)].fill = r_highlight
                    if not gov_id_iped.isnumeric():
                        ws_uasys['D' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['E']:
            try:
                ws_uasys['E' + str(cell.row)].value = str(cell.value).upper()
                if cell.value is None:
                    ws_uasys['E' + str(cell.row)].fill = y_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['F']:
            try:
                ws_uasys['F' + str(cell.row)].value = str(cell.value).upper()
                if cell.value is None:
                    ws_uasys['F' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['G']:
            try:
                if cell.value is None:
                    ws_uasys['G' + str(cell.row)].value = "N/A"
            except:
                print('Error with cell')
        for cell in ws_uasys['H']:
            try:
                if cell.value is None:
                    ws_uasys['H' + str(cell.row)].value = "N/A"
            except:
                print('Error with cell')
        for cell in ws_uasys['I']:
            try:
                ws_uasys['I' + str(cell.row)].value = str(cell.value).upper()
                if cell.value is None:
                    ws_uasys['I' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['J']:
            try:
                if cell.row >= 3:
                    region = str(cell.value)
                    if len(region) != 2:
                        ws_uasys['J' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['K']:
            try:
                if cell.value is None:
                    ws_uasys['K' + str(cell.row)].value = 'USA'
            except:
                print('Error with cell')
        for cell in ws_uasys['L']:
            try:
                if cell.value is None:
                    ws_uasys['L' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['M']:
            try:
                if cell.value is None:
                    ws_uasys['M' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['P']:
            try:
                if cell.value is None:
                    ws_uasys['P' + str(cell.row)].value = "N/A"
            except:
                print('Error with cell')
        wb_uasys.save(raw_file)
        print('Done!')

    @classmethod
    def reconcile_campuslocation(cls, wb_uasys, ws_uasys, raw_file, abbrev, ws_data_grab, ws_nces_grab):
        # If CAMPUS_LOCATION_ID is blank then assign the cell AutoGen
        for cell in ws_uasys['AK']:
            try:
                if cell.value is None:
                    ws_uasys['AK' + str(cell.row)].value = "AutoGen"
            except AttributeError:
                print('Cell is read only!')
            except TypeError:
                print('Cell is read only!')
            except:
                print('Unknown error')
        # Grabbing the missing cells in CAMP_OFFICIAL_INSTITUTION_NAME from PRIMARY_INSTITUTION_NAME
        for cell in ws_uasys['AP']:
            if cell.value is None:
                try:
                    CAMP_OFFICIAL_INSTITUTION_NAME = ws_uasys['U' + str(cell.row)].value
                    ws_uasys['AP' + str(cell.row)].value = CAMP_OFFICIAL_INSTITUTION_NAME.upper()
                except AttributeError:
                    print('NoneType object has no attribute upper')
        # for column AQ find the additional location for column AP and address for the location
        for cell in ws_uasys['AQ']:
            if cell.row >= 3:
                try:
                    lookup_institution = ws_uasys['AP' + str(cell.row)].value
                    for grab in ws_data_grab['E']:
                        if grab.value.upper == lookup_institution.upper:
                            additional_location = ws_data_grab['D' + str(grab.row)].value
                            address_additionalLocation = ws_data_grab['H' + str(grab.row)].value
                            prev_additional_location = ws_uasys['AQ' + str(cell.row - 1)].value
                            if additional_location.upper != prev_additonal_location.upper: 
                                ws_uasys['AQ' + str(cell.row)].value = str(additional_location).upper
                                ws_uasys['AS' + str(cell.row)].value = str(address_additionalLocation).upper
                                wb_uasys.save(raw_file)
                except AttributeError:
                    print('Cell is read only!')
                except TypeError:
                    print('Cell is read only!')
                except:
                    print('Unknown error')
        # Get CAMP_CAMPUS_NAME CAMP_OPED_ID and CAMP_IPED_ID from LocationName OpeId and IpedsUnitIds
        for cell in ws_uasys['AQ']:
            organization_name = str(cell.value)
            print("----------------------------------")
            print('Populating ' + organization_name + ' fields.....')
            for grab in ws_data_grab['D']:
                location_name = str(grab.value)
                if location_name.upper() == organization_name.upper():
                    CAMP_DAPID = str(ws_data_grab['A' + str(grab.row)].value)
                    CAMP_OPED_ID = str(ws_data_grab['B' + str(grab.row)].value)
                    CAMP_IPED_ID = str(ws_data_grab['C' + str(grab.row)].value)
                    ws_uasys['AL' + str(cell.row)].value = CAMP_DAPID
                    ws_uasys['AM' + str(cell.row)].value = CAMP_OPED_ID
                    ws_uasys['AN' + str(cell.row)].value = CAMP_IPED_ID
        # for any CAMP_CAMPUS_NAME ids that aren't populated by accred, find data in institution
        for cell in ws_uasys['AQ']:
            CAMP_CAMPUS_NAME = str(cell.value)
            COIN_dapid = ws_uasys['AL' + str(cell.row)].value
            COIN_opeid = ws_uasys['AM' + str(cell.row)].value
            COIN_ipedid = ws_uasys['AN' + str(cell.row)].value
            if COIN_dapid is None and COIN_opeid is None and COIN_ipedid is None:
                institution_name = ws_uasys['U' + str(cell.row)].value
                try:
                    if institution_name == CAMP_CAMPUS_NAME:
                        pop_dapid = str(ws_uasys['R' + str(cell.row)].value)
                        pop_opeid = str(ws_uasys['S' + str(cell.row)].value)
                        pop_ipedid = str(ws_uasys['T' + str(cell.row)].value)
                        ws_uasys['AL' + str(cell.row)].value = pop_dapid
                        ws_uasys['AM' + str(cell.row)].value = pop_opeid
                        ws_uasys['AN' + str(cell.row)].value = pop_ipedid
                except AttributeError:
                    print('NoneType object has no attribute')
        # If Campus dapid, opeid, ipedid is none then assign the cell NULL
        for cell in ws_uasys['AL']:
            opeid = ws_uasys['AM' + str(cell.row)].value
            ipedid = ws_uasys['AN' + str(cell.row)].value
            try:
                if cell.value is None:
                    ws_uasys['AL' + str(cell.row)].value = "NULL"
                if opeid is None:
                    ws_uasys['AM' + str(cell.row)].value = "NULL"
                if ipedid is None:
                    ws_uasys['AN' + str(cell.row)].value = "NULL"
            except AttributeError:
                print('Cell is read only!')
            except TypeError:
                print('Cell is read only!')
            except:
                print('Unknown error')
        # Get CAMP_PO_BOX_LINE and CAMP_PhoneNumberFull from CAMP_CAMPUS_NAME against LocationName fields
        for cell in ws_uasys['AQ']:
            organization_name = str(cell.value)
            for grab in ws_data_grab['D']:
                location_name = str(grab.value)
                if location_name.upper() == organization_name.upper():
                    CAMP_PhoneNumberFull = str(ws_data_grab['I' + str(grab.row)].value)
                    address_grab = str(ws_data_grab['H' + str(grab.row)].value)
                    address_grab.split(', ')
                    try:
                        if len(address_grab.split(', ')) == 1:
                            address_grab = address_grab + ", N/A, N/A, N/A, N/A, N/A, N/A"
                        if len(address_grab.split(', ')) == 2:
                            address_grab = address_grab + ", N/A, N/A, N/A, N/A, N/A"
                        if len(address_grab.split(', ')) == 2:
                            address_grab = address_grab + ", N/A, N/A, N/A, N/A, N/A"
                        if len(address_grab.split(', ')) == 3:
                            address_grab = address_grab + ", N/A, N/A, N/A, N/A"
                        if len(address_grab.split(', ')) == 4:
                            address_grab = address_grab + ", N/A, N/A, N/A"

                        GOV_ADDRESS_LINE_1, temp_LINE_2, temp_POBOX, temp_MUNI, temp_PCODE, temp1_Unknown, temp2_Unknown = address_grab.split(
                            ', ')

                        if GOV_ADDRESS_LINE_1.startswith('P.O. Box'):
                            temp_PCODE = temp_POBOX
                            temp_POBOX = GOV_ADDRESS_LINE_1
                            GOV_ADDRESS_LINE_1 = 'N/A'

                        if temp_POBOX.startswith('K'):
                            temp_POBOX = 'N/A'
                            temp_MUNI = temp_PCODE
                            temp_PCODE = temp1_Unknown

                        if temp_PCODE.startswith('P.O BOX'):
                            temp_POBOX = temp_PCODE
                            temp_PCODE = 'NULL'

                        if not temp_LINE_2.startswith('Suite'):
                            temp_MUNI = temp_LINE_2
                            temp_LINE_2 = 'N/A'
                            temp_PCODE = temp_POBOX
                            temp_POBOX = 'N/A'

                        CAMP_ADDRESS_LINE_2 = temp_LINE_2.upper()
                        CAMP_PO_BOX_LINE = temp_POBOX.strip('.')
                        CAMP_MUNICIPALITY = temp_MUNI.upper()
                        CAMP_POSTAL_CODE = temp_PCODE.strip(abbrev)

                        ws_uasys['AT' + str(cell.row)].value = CAMP_ADDRESS_LINE_2
                        ws_uasys['AU' + str(cell.row)].value = CAMP_PO_BOX_LINE
                        ws_uasys['AV' + str(cell.row)].value = CAMP_MUNICIPALITY
                        ws_uasys['AY' + str(cell.row)].value = CAMP_POSTAL_CODE
                    except ValueError:
                        ws_uasys['AT' + str(cell.row)].value = 'NULL'
                        ws_uasys['AU' + str(cell.row)].value = 'NULL'
                        ws_uasys['AV' + str(cell.row)].value = 'NULL'
                        ws_uasys['AY' + str(cell.row)].value = 'NULL'
                    except TypeError:
                        ws_uasys['AT' + str(cell.row)].value = 'NULL'
                        ws_uasys['AU' + str(cell.row)].value = 'NULL'
                        ws_uasys['AV' + str(cell.row)].value = 'NULL'
                        ws_uasys['AY' + str(cell.row)].value = 'NULL'
                    except:
                        ws_uasys['AT' + str(cell.row)].value = 'NULL'
                        ws_uasys['AU' + str(cell.row)].value = 'NULL'
                        ws_uasys['AV' + str(cell.row)].value = 'NULL'
                        ws_uasys['AY' + str(cell.row)].value = 'NULL'

                    ws_uasys['AZ' + str(cell.row)].value = CAMP_PhoneNumberFull
        # Grabbing location data from Institution section to bring it to campus/location section for main campuses/one location
        for cell in ws_uasys['AQ']:
            campus_institution = str(cell.value)
            official_institution = ws_uasys['AP' + str(cell.row)].value
            if campus_institution == official_institution and ws_uasys['AS' + str(cell.row)].value is None:
                ADDRESS_LINE_1 = ws_uasys['V' + str(cell.row)].value
                ADDRESS_LINE_2 = ws_uasys['W' + str(cell.row)].value
                PO_BOX_LINE = ws_uasys['X' + str(cell.row)].value
                MUNICIPALITY = ws_uasys['Y' + str(cell.row)].value
                STATE_REGION_SHORT = ws_uasys['Z' + str(cell.row)].value
                POSTAL_CODE = ws_uasys['AB' + str(cell.row)].value
                PhoneNumberFull = ws_uasys['AC' + str(cell.row)].value
                ws_uasys['AQ' + str(cell.row)].value = 'MAIN CAMPUS'
                ws_uasys['AS' + str(cell.row)].value = ADDRESS_LINE_1
                ws_uasys['AT' + str(cell.row)].value = ADDRESS_LINE_2
                ws_uasys['AU' + str(cell.row)].value = PO_BOX_LINE
                ws_uasys['AV' + str(cell.row)].value = MUNICIPALITY
                ws_uasys['AW' + str(cell.row)].value = STATE_REGION_SHORT
                ws_uasys['AX' + str(cell.row)].value = 'USA'
                ws_uasys['AY' + str(cell.row)].value = POSTAL_CODE
                ws_uasys['AZ' + str(cell.row)].value = PhoneNumberFull
        # Move/delete substrings from CAMP_ADDRESS_LINE_1 and moving them into respective column row
        for cell in ws_uasys['AS']:
            governing_address = str(cell.value).split()
            for index in range(len(governing_address)):
                word = governing_address[index]
                if word == 'Ste' or word == 'Ste.' or word == 'STE' or word == 'STE.' or word == 'Unit' or word == 'PO' or word == 'Suite':
                    GOV_ADDRESS_LINE_2 = str(' '.join(governing_address[index:len(governing_address)]))
                    found_pobox = GOV_ADDRESS_LINE_2.find('PO Box')
                    if found_pobox == -1:
                        ws_uasys['AU' + str(cell.row)].value = GOV_ADDRESS_LINE_2.upper()
                    else:
                        ws_uasys['AT' + str(cell.row)].value = GOV_ADDRESS_LINE_2
                        ws_uasys['AU' + str(cell.row)].value = 'N/A'
                    ADDRESS_LINE_1 = str(cell.value)
                    phrase_removal = ADDRESS_LINE_1.find(GOV_ADDRESS_LINE_2)
                    if phrase_removal != -1:
                        ws_uasys['AS' + str(cell.row)].value = ADDRESS_LINE_1.strip(GOV_ADDRESS_LINE_2)
                elif word == 'Floor' or word == 'Fl':
                    floor_num = index - 1
                    GOV_ADDRESS_LINE_2 = str(' '.join(governing_address[floor_num:len(governing_address)]))
                    ws_uasys['AT' + str(cell.row)].value = GOV_ADDRESS_LINE_2.upper()
                    ADDRESS_LINE_1 = str(cell.value)
                    phrase_removal = ADDRESS_LINE_1.find(GOV_ADDRESS_LINE_2)
                    if phrase_removal != -1:
                        ws_uasys['AS' + str(cell.row)].value = ADDRESS_LINE_1.strip(GOV_ADDRESS_LINE_2)
        # Move/delete substrings from CAMP_ADDRESS_LINE_2,
        for cell in ws_uasys['AU']:
            CAMP_PO_BOX_LINE = str(cell.value).split()
            word = CAMP_PO_BOX_LINE[0]
            if word != 'PO' or word != 'N/A':
                if word.find('STE') == -1:
                    try:
                        ADDRESS_LINE_2 = ws_uasys['AT' + str(cell.row)].value
                        ws_uasys['AU' + str(cell.row)].value = ADDRESS_LINE_2
                        ws_uasys['AT' + str(cell.row)].value = 'N/A'
                        ws_uasys['AT2'].value = 'CAMP_ADDRESS_LINE_2'
                    except AttributeError:
                        print('MergedCell object attribute value is read-only')
                else:
                    CAMP_POSTAL_CODE = ws_uasys['AV' + str(cell.row)].value
                    CAMP_MUNICIPALITY = ws_uasys['AU' + str(cell.row)].value
                    ws_uasys['AY' + str(cell.row)].value = CAMP_POSTAL_CODE
                    ws_uasys['AV' + str(cell.row)].value = CAMP_MUNICIPALITY
                    ws_uasys['AU' + str(cell.row)].value = 'N/A'
        for cell in ws_uasys['AY']:
            POSTAL_CODE = str(cell.value)
            try:
                if POSTAL_CODE.isalpha() and POSTAL_CODE != 'N/A':
                    CAMP_MUNICIPALITY = ws_uasys['AY' + str(cell.row)].value
                    CAMP_ADDRESS_LINE_2 = ws_uasys['AV' + str(cell.row)].value
                    ws_uasys['AT' + str(cell.row)].value = CAMP_ADDRESS_LINE_2
                    ws_uasys['AV' + str(cell.row)].value = CAMP_MUNICIPALITY
                    ws_uasys['AY' + str(cell.row)].value = ''
                if cell.value == 'N/A' or cell.value == 'N/':
                    CAMP_POSTAL_CODE = ws_uasys['AV' + str(cell.row)].value
                    ADDRESS_LINE_2 = ws_uasys['AU' + str(cell.row)].value
                    ws_uasys['AY' + str(cell.row)].value = CAMP_POSTAL_CODE
                    ws_uasys['AT' + str(cell.row)].value = ADDRESS_LINE_2
                    ws_uasys['AU' + str(cell.row)].value = 'N/A'
                    ws_uasys['AU2'].value = 'CAMP_PO_BOX_LINE'
                    ws_uasys['AV' + str(cell.row)].value = ''
                postal_code_list = str(cell.value).split()
                word = postal_code_list[0]
                if word.isalpha() and len(word) <= 2:
                    STATE_REGION_SHORT = str(word).strip('[]')
                    ws_uasys['AW' + str(cell.row)].value = STATE_REGION_SHORT
                    ws_uasys['AY' + str(cell.row)].value = str(cell.value).strip(STATE_REGION_SHORT)
            except IndexError:
                print('list index out of range')
            except AttributeError:
                print('MergedCell object attribute value is read-only')
        # Check to see if campus is inactive/closed according to NCES database
        for cell in ws_uasys['AQ']:
            organization_name = str(cell.value)
            for look in ws_nces_grab['B']:
                nces_institution = str(look.value)
                if nces_institution.upper() == organization_name.upper():
                    institution_closed = ws_nces_grab['W' + str(look.row)].value
                    found_two = str(institution_closed).find('-2')
                    if found_two < 0:
                        ws_uasys['BD' + str(cell.row)].value = institution_closed
                        ws_uasys['BC' + str(cell.row)].value = 'Y'
        # If CAMPUS_RECORD_SOURCE is blank then assign the cell N/A
        for cell in ws_uasys['BE']:
            try:
                if cell.value is None:
                    ws_uasys['BE' + str(cell.row)].value = "N/A"
            except AttributeError:
                print('Cell is read only!')
            except TypeError:
                print('Cell is read only!')
            except:
                print('Unknown error')
        print('Done!')
        wb_uasys.save(raw_file)

    @classmethod
    def clean_campuslocation(cls, wb_uasys, ws_uasys, raw_file):
        for cell in ws_uasys['AK']:
            try:
                if cell.value is None:
                    ws_uasys['AK' + str(cell.row)].value = 'AutoGen'
            except:
                print('Error with cell')
        for cell in ws_uasys['AL']:
            try:
                if cell.value is None:
                    ws_uasys['AL' + str(cell.row)].value = "NULL"
            except:
                print('Error with cell')
        for cell in ws_uasys['AM']:
            try:
                if cell.value is None:
                    ws_uasys['AM' + str(cell.row)].value = "NULL"
            except:
                print('Error with cell')
        for cell in ws_uasys['AN']:
            try:
                if cell.value is None:
                    ws_uasys['AN' + str(cell.row)].value = "NULL"
            except:
                print('Error with cell')
        yellow = 'FFFF00'
        red = 'FF6666'
        y_highlight = PatternFill(patternType='solid', fgColor=yellow)
        r_highlight = PatternFill(patternType='solid', fgColor=red)
        for cell in ws_uasys['AL']:
            try:
                if cell.row >= 3:
                    gov_id = str(cell.value)
                    gov_id_oped = str(ws_uasys['AM' + str(cell.row)].value)
                    gov_id_iped = str(ws_uasys['AN' + str(cell.row)].value)
                    if not gov_id.isnumeric():
                        ws_uasys['AL' + str(cell.row)].fill = r_highlight
                    if not gov_id_oped.isnumeric():
                        ws_uasys['AM' + str(cell.row)].fill = r_highlight
                    if not gov_id_iped.isnumeric():
                        ws_uasys['AN' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['AP']:
            try:
                ws_uasys['AP' + str(cell.row)].value = str(cell.value).upper()
                if cell.value is None:
                    ws_uasys['AP' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['AS']:
            try:
                ws_uasys['AS' + str(cell.row)].value = str(cell.value).upper()
                if cell.value is None:
                    ws_uasys['AS' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['AT']:
            try:
                address_one = cell.value
                if address_one is None:
                    ws_uasys['AT' + str(cell.row)].value = 'N/A'
                if address_one.find('PO') != -1:
                    ws_uasys['AT' + str(cell.row)].fill = y_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['AU']:
            if cell.row >= 3:
                try:
                    address_two = str(cell.value)
                    if cell.value is None:
                        ws_uasys['AU' + str(cell.row)].value = 'N/A'
                        wb_uasys.save(raw_file)
                    if address_two.find('PO') == -1 and address_two != 'N/A':
                        ws_uasys['AU' + str(cell.row)].fill = y_highlight
                except:
                    print('Error with cell')
        for cell in ws_uasys['AV']:
            try:
                ws_uasys['AV' + str(cell.row)].value = str(cell.value).upper()
                if cell.value is None:
                    ws_uasys['AV' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['AW']:
            try:
                if cell.row >= 3:
                    region = str(cell.value)
                    if len(region) != 2:
                        ws_uasys['AW' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['AX']:
            try:
                if cell.value is None:
                    ws_uasys['AX' + str(cell.row)].value = 'USA'
            except:
                print('Error with cell')
        for cell in ws_uasys['AY']:
            try:
                if cell.value is None:
                    ws_uasys['AY' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['AZ']:
            try:
                if cell.value is None:
                    ws_uasys['AZ' + str(cell.row)].fill = r_highlight
            except:
                print('Error with cell')
        for cell in ws_uasys['BA']:
            if cell.row >= 3:
                try:
                    if cell.value == 'Manually Find' or cell.value is None:
                        ws_uasys['BA' + str(cell.row)].fill = y_highlight
                except:
                    print('Error with cell')
        for cell in ws_uasys['BB']:
            if cell.row >= 3:
                try:
                    if cell.value == 'Manually Find' or cell.value is None:
                        ws_uasys['BB' + str(cell.row)].fill = y_highlight
                except:
                    print('Error with cell')
        for cell in ws_uasys['BE']:
            try:
                if cell.value is None:
                    ws_uasys['BE' + str(cell.row)].value = "N/A"
            except:
                print('Error with cell')
        wb_uasys.save(raw_file)
        print('Done!')

    @classmethod
    def ai_campuslocation(cls, wb_uasys, ws_uasys, raw_file):
        max_row = ws_uasys.max_row
        for cell in ws_uasys['AP']:
            progress = cell.row/max_row
            sys.stdout.write('\r')
            sys.stdout.write("[%-20s] %d%%" % ('=' * int(max_row * progress), float(cell.row/max_row)*100))
            try:
                if cell.row >= 3:
                    cell_prev = int(cell.row) - 1
                    institution_name = str(cell.value)
                    municipality = str(ws_uasys['AV' + str(cell.row)].value)
                    state = str(ws_uasys['AW' + str(cell.row)].value)
                    if institution_name != ws_uasys['AP' + str(cell_prev)].value and ws_uasys[
                        'BA' + str(cell.row)].value is None:
                        API_KEY = open(r"C:\Users\Wayne\Work Stuff\Data Conversion\API Key.txt").read()
                        openai.api_key = API_KEY
                        response = openai.ChatCompletion.create(
                            model="gpt-3.5-turbo",
                            messages=[
                                {"role": "system", "content": "You are a data analyst reconciling missing data."},
                                {"role": "user",
                                 "content": "Don't include the question in your response, what is the date when"
                                            "Texas State University at San Marcos, TX founded?"},
                                {"role": "assistant", "content": "1899-01-01"},
                                {"role": "user",
                                 "content": "Don't include the question in your response, what is the date when"
                                            "SAINT MARY'S COLLEGE OF CALIFORNIA at MORAGA, CA founded?"},
                                {"role": "assistant", "content": "1863-01-01"},
                                {"role": "user",
                                 "content": "Don't include the question in your response, what is the date when "
                                            + institution_name + ' at ' + municipality + ', ' + state + " founded?"}
                            ]
                        )
                        reply_content = response.choices[0].message.content
                        if DataFile.has_numbers(reply_content):
                            ws_uasys['BA' + str(cell.row)].value = str(reply_content)
                        else:
                            ws_uasys['BA' + str(cell.row)].value = 'NULL'
                        wb_uasys.save(raw_file)
                        time.sleep(1)
                wb_uasys.save(raw_file)
            except openai.error.APIError:
                print('Bad Gateway')
            except openai.error.ServiceUnavailableError:
                print('Server overload')
            except TimeoutError:
                print('Read Operation Timeout')
            except:
                print('Server Overload')
            sys.stdout.flush()

        for cell in ws_uasys['AP']:
            progress = cell.row/max_row
            sys.stdout.write('\r')
            sys.stdout.write("[%-20s] %d%%" % ('=' * int(max_row * progress), float(cell.row/max_row)*100))
            try:
                if cell.row >= 3:
                    cell_prev = int(cell.row) - 1
                    institution_name = str(cell.value)
                    municipality = str(ws_uasys['AV' + str(cell.row)].value)
                    state = str(ws_uasys['AW' + str(cell.row)].value)
                    if institution_name != ws_uasys['AP' + str(cell_prev)].value and ws_uasys[
                        'BB' + str(cell.row)].value is None:
                        API_KEY = open(r"C:\Users\Wayne\Work Stuff\Data Conversion\API Key.txt").read()
                        openai.api_key = API_KEY
                        response = openai.ChatCompletion.create(
                            model="gpt-3.5-turbo",
                            messages=[
                                {"role": "system", "content": "You are a data analyst reconciling missing data."},
                                {"role": "user",
                                 "content": "Don't include the question in your response, When was this "
                                            "campus named Texas State University in San Marcos, TX?"},
                                {"role": "assistant", "content": "2013-01-01"},
                                {"role": "user",
                                 "content": "Don't include the question in your response, When was this "
                                            "campus named SAINT MARY'S COLLEGE OF CALIFORNIA in MORAGA, CA?"},
                                {"role": "assistant", "content": "1863-01-01"},
                                {"role": "user", "content": "If you can not find the date please respond with N/A."},
                                {"role": "assistant", "content": "N/A"},
                                {"role": "user",
                                 "content": "Don't include the question in your response, When was this "
                                            "campus named " + institution_name + ' in ' + municipality + ', '
                                            + state + "?"}
                            ]
                        )
                        reply_content = response.choices[0].message.content
                        if DataFile.has_numbers(reply_content):
                            ws_uasys['BB' + str(cell.row)].value = str(reply_content)
                        else:
                            ws_uasys['BB' + str(cell.row)].value = 'NULL'
                        wb_uasys.save(raw_file)
                        time.sleep(1)
                wb_uasys.save(raw_file)
            except openai.error.APIError:
                print('Bad Gateway')
            except openai.error.ServiceUnavailableError:
                print('Server overload')
            except TimeoutError:
                print('Read Operation Timeout')
            except:
                print('Server Overload')
            sys.stdout.flush()
