from openpyxl import workbook, load_workbook

raw_file = input("File location in explorer(.xlsx): ")
wrong_input = raw_file.find(".xlsx")
if wrong_input == -1:
    print("Be sure to add .xlsx to end of file location!")
    raw_file = input("File location is explorer(.xlsx)")
wb_uasys = load_workbook(raw_file)
wb_data_grab = load_workbook("AccreditationData.xlsx")
wb_nces_grab = load_workbook("Data_3-14-2023---623.xlsx")
sheet_name = input("Name of sheet in raw file: ")
ws_uasys = wb_uasys[sheet_name]
ws_data_grab = wb_data_grab["InstituteCampuses"]
ws_nces_grab = wb_nces_grab["Data_3-14-2023---623"]
abbrev = input("State Abbreviation of worksheet is needed: ")
# If GOVERNING_ORGANIZATION_ID is blank then assign the cell AutoGen
for cell in ws_uasys['A']:
    try:
        if cell.value is None:
            ws_uasys['A' + str(cell.row)].value = "AutoGen"
    except AttributeError:
        print(cell + ' is read only!')
    except TypeError:
        print('Cell is read only!')
    except:
        print('Unknown error')
# Get primary institution name and compare it against cells in additional sites location name,
# if match: access Parent
# Name Cell and return cell data to populate Governing Org name of same row as primary institution name
for cell in ws_uasys['U']:
    institute_name = str(cell.value)
    print("----------------------------------")
    print(cell.value)

    for grab in ws_data_grab['D']:
        location_name = str(grab.value)

        if location_name.upper() == institute_name.upper():
            char = 'D'
            p_char = chr(ord(char) + 1)
            parent_name = str(ws_data_grab[p_char + str(grab.row)].value)

            if parent_name == "-":
                parent_name = str(ws_data_grab['D' + str(grab.row)].value)

            ws_uasys['E' + str(cell.row)].value = parent_name
            print(ws_uasys['E' + str(cell.row)].value)
            print("----------------------------------")
print("Populating associated fields.....hold on.....")
# Get Governing_Organization_Name's DAPIP, OPE, and IPEDSID IDs from data_grab
for cell in ws_uasys['E']:
    institution_govern = str(cell.value)

    for grab in ws_data_grab['D']:
        location_name = str(grab.value)

        if location_name.upper() == institution_govern.upper():
            GOV_DAPID = str(ws_data_grab['A' + str(grab.row)].value)
            GOV_OPEID = str(ws_data_grab['B' + str(grab.row)].value)
            GOV_IPEDID = str(ws_data_grab['C' + str(grab.row)].value)

            ws_uasys['B' + str(cell.row)].value = GOV_DAPID
            ws_uasys['C' + str(cell.row)].value = GOV_OPEID
            ws_uasys['D' + str(cell.row)].value = GOV_IPEDID
# Get GOV address line 1, GOV_MUNICIPALITY, GOV postal code
for cell in ws_uasys['E']:
    institution_govern = str(cell.value)

    for grab in ws_data_grab['D']:
        location_name = str(grab.value)
        if location_name.upper() == institution_govern.upper():
            address_grab = str(ws_data_grab['H' + str(grab.row)].value)
            address_grab.split(', ')
            try:
                if len(address_grab.split(', ')) == 1:
                    address_grab = address_grab + ", N/A, N/A, N/A, N/A, N/A, N/A"
                if len(address_grab.split(', ')) == 2:
                    address_grab = address_grab + ", N/A, N/A, N/A, N/A, N/A"
                if len(address_grab.split(', ')) == 2:
                    address_grab = address_grab + ", N/A, N/A, N/A, N/A, N/A"
                if len(address_grab.split(', ')) == 3:
                    address_grab = address_grab + ", N/A, N/A, N/A, N/A"
                if len(address_grab.split(', ')) == 4:
                    address_grab = address_grab + ", N/A, N/A, N/A"

                GOV_ADDRESS_LINE_1, temp_LINE_2, temp_POBOX, temp_MUNI, temp_PCODE, temp1_Unknown, temp2_Unknown = address_grab.split(
                    ', ')

                if GOV_ADDRESS_LINE_1.startswith('P.O. Box'):
                    temp_PCODE = temp_POBOX
                    temp_POBOX = GOV_ADDRESS_LINE_1
                    GOV_ADDRESS_LINE_1 = 'N/A'

                if temp_POBOX.startswith('K'):
                    temp_POBOX = 'N/A'
                    temp_MUNI = temp_PCODE
                    temp_PCODE = temp1_Unknown

                if temp_PCODE.startswith('P.O BOX'):
                    temp_POBOX = temp_PCODE
                    temp_PCODE = 'NULL'

                if not temp_LINE_2.startswith('Suite'):
                    temp_MUNI = temp_LINE_2
                    temp_LINE_2 = 'N/A'
                    temp_PCODE = temp_POBOX
                    temp_POBOX = 'N/A'
                if temp_MUNI.startswith(abbrev.upper()):
                    temp_PCODE = temp_MUNI
                    temp_MUNI = temp_POBOX
                    temp_POBOX = 'N/A'

                GOV_ADDRESS_LINE_2 = temp_LINE_2.upper()
                GOV_PO_BOX_LINE = temp_POBOX.strip('.')
                GOV_MUNICIPALITY = temp_MUNI.upper()
                GOV_POSTAL_CODE = temp_PCODE.strip(abbrev.upper())

                ws_uasys['F' + str(cell.row)].value = GOV_ADDRESS_LINE_1
                ws_uasys['G' + str(cell.row)].value = GOV_ADDRESS_LINE_2
                ws_uasys['H' + str(cell.row)].value = GOV_PO_BOX_LINE
                ws_uasys['I' + str(cell.row)].value = GOV_MUNICIPALITY
                ws_uasys['L' + str(cell.row)].value = GOV_POSTAL_CODE
            except ValueError:
                ws_uasys['F' + str(cell.row)].value = 'NULL'
                ws_uasys['G' + str(cell.row)].value = 'NULL'
                ws_uasys['H' + str(cell.row)].value = 'NULL'
                ws_uasys['I' + str(cell.row)].value = 'NULL'
                ws_uasys['L' + str(cell.row)].value = 'NULL'
            except:
                ws_uasys['F' + str(cell.row)].value = 'NULL'
                ws_uasys['G' + str(cell.row)].value = 'NULL'
                ws_uasys['H' + str(cell.row)].value = 'NULL'
                ws_uasys['I' + str(cell.row)].value = 'NULL'
                ws_uasys['L' + str(cell.row)].value = 'NULL'
# If GOV_STATE_REGION_SHORT is blank then assign worksheet state
for cell in ws_uasys['J']:
    try:
        if cell.value is None:
            ws_uasys['J' + str(cell.row)].value = abbrev.upper()
    except AttributeError:
        print('Cell is read only!')
    except TypeError:
        print('Cell is read only!')
    except:
        print('Unknown error')
# If GOV_COUNTRY_CODE is blank then assign USA
for cell in ws_uasys['K']:
    try:
        if cell.value is None:
            ws_uasys['K' + str(cell.row)].value = "USA"
    except AttributeError:
        print('Cell is read only!')
    except TypeError:
        print('Cell is read only!')
    except:
        print('Unknown error')
# Get GOV_PhoneNumberFull
for cell in ws_uasys['E']:
    institution_govern = str(cell.value)

    for grab in ws_data_grab['D']:
        location_name = str(grab.value)
        if location_name.upper() == institution_govern.upper():
            phoneNumber_grab = str(ws_data_grab['I' + str(grab.row)].value)
            ws_uasys['M' + str(cell.row)].value = phoneNumber_grab

            if ws_uasys['M' + str(cell.row)].value is None:
                print('No phone number from Accreditation Database : Searching')
                for look in ws_nces_grab['B']:
                    nces_institution = str(look.value)
                    if nces_institution.upper() == institution_govern.upper():
                        phoneNumber_grab = str(ws_nces_grab['L' + str(grab.row)].value)
                        ws_uasys['M' + str(cell.row)].value = phoneNumber_grab
# Check to see if GOV_ORG is inactive/closed according to NCES database
for cell in ws_uasys['E']:
    institution_govern = str(cell.value)
    for look in ws_nces_grab['B']:
        nces_institution = str(look.value)
        if nces_institution.upper() == institution_govern.upper():
            institution_closed = ws_nces_grab['W' + str(look.row)].value
            found_two = str(institution_closed).find('-2')
            if found_two < 0:
                ws_uasys['O' + str(cell.row)].value = institution_closed
# If GOV_RECORD_SOURCE is blank then assign the cell N/A
for cell in ws_uasys['P']:
    try:
        if cell.value is None:
            ws_uasys['P' + str(cell.row)].value = "N/A"
    except AttributeError:
        print('Cell is read only!')
    except TypeError:
        print('Cell is read only!')
    except:
        print('Unknown error')
# if not in data_grab then search nces_grab database
for cell in ws_uasys['E']:
    try:
        if cell.value is None:
            search_institution = ws_uasys['U' + str(cell.row)].value
            for look in ws_nces_grab['B']:
                nces_institution = str(look.value)
                if nces_institution.upper() == search_institution.upper():
                    GOV_DAPID = ws_uasys['R' + str(cell.row)].value
                    GOV_OPEID = ws_uasys['S' + str(cell.row)].value
                    GOV_IPEDID = ws_uasys['T' + str(cell.row)].value
                    PRIMARY_INSTITUTION_NAME = nces_institution.upper()
                    GOV_ADDRESS_LINE_1 = ws_nces_grab['I' + str(look.row)].value
                    GOV_MUNICIPALITY = ws_nces_grab['J' + str(look.row)].value
                    GOV_STATE_REGION_SHORT = ws_nces_grab['C' + str(look.row)].value
                    GOV_POSTAL_CODE = ws_nces_grab['K' + str(look.row)].value
                    GOV_PhoneNumberFull = ws_nces_grab['L' + str(look.row)].value

                    ws_uasys['B' + str(cell.row)].value = GOV_DAPID
                    ws_uasys['C' + str(cell.row)].value = GOV_OPEID
                    ws_uasys['D' + str(cell.row)].value = GOV_IPEDID
                    ws_uasys['E' + str(cell.row)].value = PRIMARY_INSTITUTION_NAME.upper()
                    ws_uasys['F' + str(cell.row)].value = GOV_ADDRESS_LINE_1
                    ws_uasys['I' + str(cell.row)].value = GOV_MUNICIPALITY
                    ws_uasys['J' + str(cell.row)].value = GOV_STATE_REGION_SHORT
                    ws_uasys['L' + str(cell.row)].value = GOV_POSTAL_CODE
                    ws_uasys['M' + str(cell.row)].value = GOV_PhoneNumberFull
    except AttributeError:
        print('Cell is read only!')
    except TypeError:
        print('Cell is read only!')
    except:
        print('Unknown error')
# Move/delete substrings from GOV_ADDRESS_LINE_1 and moving them into respective column row
for cell in ws_uasys['F']:
    governing_address = str(cell.value).split()
    for index in range(len(governing_address)):
        word = governing_address[index]
        if word == 'Ste' or word == 'Ste.' or word == 'Unit' or word == 'PO' or word == 'Suite':
            GOV_ADDRESS_LINE_2 = str(' '.join(governing_address[index:len(governing_address)]))
            found_pobox = GOV_ADDRESS_LINE_2.find('PO Box')
            if found_pobox == -1:
                ws_uasys['G' + str(cell.row)].value = GOV_ADDRESS_LINE_2.upper()
            else:
                ws_uasys['H' + str(cell.row)].value = GOV_ADDRESS_LINE_2
                ws_uasys['G' + str(cell.row)].value = 'N/A'
            ADDRESS_LINE_1 = str(cell.value)
            phrase_removal = ADDRESS_LINE_1.find(GOV_ADDRESS_LINE_2)
            if phrase_removal != -1:
                ws_uasys['F' + str(cell.row)].value = ADDRESS_LINE_1.strip(GOV_ADDRESS_LINE_2)
        elif word == 'Floor' or word == 'Fl':
            floor_num = index - 1
            GOV_ADDRESS_LINE_2 = str(' '.join(governing_address[floor_num:len(governing_address)]))
            ws_uasys['G' + str(cell.row)].value = GOV_ADDRESS_LINE_2.upper()
            ADDRESS_LINE_1 = str(cell.value)
            phrase_removal = ADDRESS_LINE_1.find(GOV_ADDRESS_LINE_2)
            if phrase_removal != -1:
                ws_uasys['F' + str(cell.row)].value = ADDRESS_LINE_1.strip(GOV_ADDRESS_LINE_2)
# Move/delete substrings from GOV_POSTAL_CODE to GOV_MUNICIPALITY, GOV_MUNICIPALITY moves to GOV_ADDRESS_LINE_2
for cell in ws_uasys['L']:
    postal_code = str(cell.value).split()
    for index in range(len(postal_code)):
        word = postal_code[index]
        if not word.isalpha():
            continue
        else:
            GOV_POSTAL_CODE = str(' '.join(postal_code[index:len(postal_code)]))
            if GOV_POSTAL_CODE.isalpha():
                try:
                    GOV_ADDRESS_LINE_1 = ws_uasys['I' + str(cell.row)].value
                    GOV_MUNICIPALITY = ws_uasys['L' + str(cell.row)].value
                    ws_uasys['I' + str(cell.row)].value = GOV_MUNICIPALITY
                    ws_uasys['G' + str(cell.row)].value = GOV_ADDRESS_LINE_1
                    ws_uasys['L' + str(cell.row)].value = ''
                except AttributeError:
                    print('MergedCell object attribute value is read-only')
            else:
                GOV_STATE_REGION_SHORT = str(postal_code[index])
                ws_uasys['J' + str(cell.row)].value = GOV_STATE_REGION_SHORT
                ws_uasys['L' + str(cell.row)].value = GOV_POSTAL_CODE.strip(GOV_STATE_REGION_SHORT)
# Check to see if institution is inactive/closed according to NCES database
for cell in ws_uasys['E']:
    institution_govern = str(cell.value)
    cell_prev = int(cell.row) - 1
    try:
        if cell_prev != 0 and institution_govern.upper() != ws_uasys['E' + str(cell_prev)].value.upper():
            for look in ws_nces_grab['B']:
                nces_institution = str(look.value)
                if nces_institution.upper() == institution_govern.upper():
                    institution_closed = ws_nces_grab['W' + str(look.row)].value
                    found_two = str(institution_closed).find('-2')
                    if found_two < 0:
                        ws_uasys['AI' + str(cell.row)].value = institution_closed
    except AttributeError:
        print("----------------------------------")
        print('NoneType for: ' + str(cell.value))
    except TypeError:
        print('NoneType')
    except:
        print('Unknown error')

print('Done!')
wb_uasys.save(raw_file)
