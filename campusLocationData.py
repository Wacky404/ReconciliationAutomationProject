from openpyxl import workbook, load_workbook

raw_file = input("File location in explorer(.xlsx): ")
wrong_input = raw_file.find(".xlsx")
if wrong_input == -1:
    print("Be sure to add .xlsx to end of file location!")
    raw_file = input("File location is explorer(.xlsx)")
wb_uasys = load_workbook(raw_file)
wb_data_grab = load_workbook("AccreditationData.xlsx")
wb_nces_grab = load_workbook("Data_3-14-2023---623.xlsx")
sheet_name = input("Name of sheet in raw file: ")
ws_uasys = wb_uasys[sheet_name]
ws_data_grab = wb_data_grab["InstituteCampuses"]
ws_nces_grab = wb_nces_grab["Data_3-14-2023---623"]
abbrev = input("State Abbreviation of worksheet is needed: ")
# If CAMPUS_LOCATION_ID is blank then assign the cell AutoGen
for cell in ws_uasys['AK']:
    try:
        if cell.value is None:
            ws_uasys['AK' + str(cell.row)].value = "AutoGen"
    except AttributeError:
        print('Cell is read only!')
    except TypeError:
        print('Cell is read only!')
    except:
        print('Unknown error')
# Grabbing the missing cells in CAMP_OFFICIAL_INSTITUTION_NAME from PRIMARY_INSTITUTION_NAME
for cell in ws_uasys['AP']:
    if cell.value is None:
        try:
            CAMP_OFFICIAL_INSTITUTION_NAME = ws_uasys['U' + str(cell.row)].value
            ws_uasys['AP' + str(cell.row)].value = CAMP_OFFICIAL_INSTITUTION_NAME.upper()
            ws_uasys['AQ' + str(cell.row)].value = CAMP_OFFICIAL_INSTITUTION_NAME.upper()
        except AttributeError:
            print('NoneType object has no attribute upper')
# Get CAMP_OFFICIAL_INSTITUTION_NAME CAMP_OPED_ID and CAMP_IPED_ID from LocationName OpeId and IpedsUnitIds
for cell in ws_uasys['AP']:
    organization_name = str(cell.value)
    print("----------------------------------")
    print('Populating ' + organization_name + ' fields.....')
    for grab in ws_data_grab['D']:
        location_name = str(grab.value)
        if location_name.upper() == organization_name.upper():
            CAMP_DAPID = str(ws_data_grab['A' + str(grab.row)].value)
            CAMP_OPED_ID = str(ws_data_grab['B' + str(grab.row)].value)
            CAMP_IPED_ID = str(ws_data_grab['C' + str(grab.row)].value)
            ws_uasys['AL' + str(cell.row)].value = CAMP_DAPID
            ws_uasys['AM' + str(cell.row)].value = CAMP_OPED_ID
            ws_uasys['AN' + str(cell.row)].value = CAMP_IPED_ID
# for any CAMP_OFFICIAL_INSTITUTION_NAME ids that aren't populated by accred, find data in institution
for cell in ws_uasys['AP']:
    CAMP_OFFICIAL_INSTITUTION_NAME = str(cell.value)
    COIN_dapid = ws_uasys['AL' + str(cell.row)].value
    COIN_opeid = ws_uasys['AM' + str(cell.row)].value
    COIN_ipedid = ws_uasys['AN' + str(cell.row)].value
    if COIN_dapid is None and COIN_opeid is None and COIN_ipedid is None:
        institution_name = ws_uasys['U' + str(cell.row)].value
        try:
            if institution_name == CAMP_OFFICIAL_INSTITUTION_NAME:
                pop_dapid = str(ws_uasys['R' + str(cell.row)].value)
                pop_opeid = str(ws_uasys['S' + str(cell.row)].value)
                pop_ipedid = str(ws_uasys['T' + str(cell.row)].value)
                ws_uasys['AL' + str(cell.row)].value = pop_dapid
                ws_uasys['AM' + str(cell.row)].value = pop_opeid
                ws_uasys['AN' + str(cell.row)].value = pop_ipedid
        except AttributeError:
            print('NoneType object has no attribute')
# # Get CAMP_PO_BOX_LINE and CAMP_PhoneNumberFull from CAMP_OFFICIAL_INSTITUTION_NAME against LocationName fields
# for cell in ws_uasys['AP']:
#     organization_name = str(cell.value)
#     for grab in ws_data_grab['D']:
#         location_name = str(grab.value)
#         if location_name.upper() == organization_name.upper():
#             CAMP_PhoneNumberFull = str(ws_data_grab['I' + str(grab.row)].value)
#             address_grab = str(ws_data_grab['H' + str(grab.row)].value)
#             address_grab.split(', ')
#             try:
#                 if len(address_grab.split(', ')) == 1:
#                     address_grab = address_grab + ", N/A, N/A, N/A, N/A, N/A, N/A"
#                 if len(address_grab.split(', ')) == 2:
#                     address_grab = address_grab + ", N/A, N/A, N/A, N/A, N/A"
#                 if len(address_grab.split(', ')) == 2:
#                     address_grab = address_grab + ", N/A, N/A, N/A, N/A, N/A"
#                 if len(address_grab.split(', ')) == 3:
#                     address_grab = address_grab + ", N/A, N/A, N/A, N/A"
#                 if len(address_grab.split(', ')) == 4:
#                     address_grab = address_grab + ", N/A, N/A, N/A"
#
#                 GOV_ADDRESS_LINE_1, temp_LINE_2, temp_POBOX, temp_MUNI, temp_PCODE, temp1_Unknown, temp2_Unknown = address_grab.split(
#                     ', ')
#
#                 if GOV_ADDRESS_LINE_1.startswith('P.O. Box'):
#                     temp_PCODE = temp_POBOX
#                     temp_POBOX = GOV_ADDRESS_LINE_1
#                     GOV_ADDRESS_LINE_1 = 'N/A'
#
#                 if temp_POBOX.startswith('K'):
#                     temp_POBOX = 'N/A'
#                     temp_MUNI = temp_PCODE
#                     temp_PCODE = temp1_Unknown
#
#                 if temp_PCODE.startswith('P.O BOX'):
#                     temp_POBOX = temp_PCODE
#                     temp_PCODE = 'NULL'
#
#                 if not temp_LINE_2.startswith('Suite'):
#                     temp_MUNI = temp_LINE_2
#                     temp_LINE_2 = 'N/A'
#                     temp_PCODE = temp_POBOX
#                     temp_POBOX = 'N/A'
#
#                 CAMP_ADDRESS_LINE_2 = temp_LINE_2.upper()
#                 CAMP_PO_BOX_LINE = temp_POBOX.strip('.')
#                 CAMP_MUNICIPALITY = temp_MUNI.upper()
#                 CAMP_POSTAL_CODE = temp_PCODE.strip(abbrev)
#
#                 ws_uasys['AT' + str(cell.row)].value = CAMP_ADDRESS_LINE_2
#                 ws_uasys['AU' + str(cell.row)].value = CAMP_PO_BOX_LINE
#                 ws_uasys['AV' + str(cell.row)].value = CAMP_MUNICIPALITY
#                 ws_uasys['AY' + str(cell.row)].value = CAMP_POSTAL_CODE
#             except ValueError:
#                 ws_uasys['AT' + str(cell.row)].value = 'NULL'
#                 ws_uasys['AU' + str(cell.row)].value = 'NULL'
#                 ws_uasys['AV' + str(cell.row)].value = 'NULL'
#                 ws_uasys['AY' + str(cell.row)].value = 'NULL'
#             except TypeError:
#                 ws_uasys['AT' + str(cell.row)].value = 'NULL'
#                 ws_uasys['AU' + str(cell.row)].value = 'NULL'
#                 ws_uasys['AV' + str(cell.row)].value = 'NULL'
#                 ws_uasys['AY' + str(cell.row)].value = 'NULL'
#             except:
#                 ws_uasys['AT' + str(cell.row)].value = 'NULL'
#                 ws_uasys['AU' + str(cell.row)].value = 'NULL'
#                 ws_uasys['AV' + str(cell.row)].value = 'NULL'
#                 ws_uasys['AY' + str(cell.row)].value = 'NULL'
#
#             ws_uasys['AZ' + str(cell.row)].value = CAMP_PhoneNumberFull
# # Grabbing location data from Institution section to bring it to campus/location section for main campuses/one location
# for cell in ws_uasys['AQ']:
#     campus_institution = str(cell.value)
#     official_institution = ws_uasys['AP' + str(cell.row)].value
#     if campus_institution == official_institution and ws_uasys['AS' + str(cell.row)].value is None:
#         ADDRESS_LINE_1 = ws_uasys['V' + str(cell.row)].value
#         ADDRESS_LINE_2 = ws_uasys['W' + str(cell.row)].value
#         PO_BOX_LINE = ws_uasys['X' + str(cell.row)].value
#         MUNICIPALITY = ws_uasys['Y' + str(cell.row)].value
#         STATE_REGION_SHORT = ws_uasys['Z' + str(cell.row)].value
#         POSTAL_CODE = ws_uasys['AB' + str(cell.row)].value
#         PhoneNumberFull = ws_uasys['AC' + str(cell.row)].value
#         ws_uasys['AS' + str(cell.row)].value = ADDRESS_LINE_1
#         ws_uasys['AT' + str(cell.row)].value = ADDRESS_LINE_2
#         ws_uasys['AU' + str(cell.row)].value = PO_BOX_LINE
#         ws_uasys['AV' + str(cell.row)].value = MUNICIPALITY
#         ws_uasys['AW' + str(cell.row)].value = STATE_REGION_SHORT
#         ws_uasys['AX' + str(cell.row)].value = 'USA'
#         ws_uasys['AY' + str(cell.row)].value = POSTAL_CODE
#         ws_uasys['AZ' + str(cell.row)].value = PhoneNumberFull
# # Move/delete substrings from CAMP_ADDRESS_LINE_1 and moving them into respective column row
# for cell in ws_uasys['AS']:
#     governing_address = str(cell.value).split()
#     for index in range(len(governing_address)):
#         word = governing_address[index]
#         if word == 'Ste' or word == 'Ste.' or word == 'STE' or word == 'STE.' or word == 'Unit' or word == 'PO' or word == 'Suite':
#             GOV_ADDRESS_LINE_2 = str(' '.join(governing_address[index:len(governing_address)]))
#             found_pobox = GOV_ADDRESS_LINE_2.find('PO Box')
#             if found_pobox == -1:
#                 ws_uasys['AU' + str(cell.row)].value = GOV_ADDRESS_LINE_2.upper()
#             else:
#                 ws_uasys['AT' + str(cell.row)].value = GOV_ADDRESS_LINE_2
#                 ws_uasys['AU' + str(cell.row)].value = 'N/A'
#             ADDRESS_LINE_1 = str(cell.value)
#             phrase_removal = ADDRESS_LINE_1.find(GOV_ADDRESS_LINE_2)
#             if phrase_removal != -1:
#                 ws_uasys['AS' + str(cell.row)].value = ADDRESS_LINE_1.strip(GOV_ADDRESS_LINE_2)
#         elif word == 'Floor' or word == 'Fl':
#             floor_num = index - 1
#             GOV_ADDRESS_LINE_2 = str(' '.join(governing_address[floor_num:len(governing_address)]))
#             ws_uasys['AT' + str(cell.row)].value = GOV_ADDRESS_LINE_2.upper()
#             ADDRESS_LINE_1 = str(cell.value)
#             phrase_removal = ADDRESS_LINE_1.find(GOV_ADDRESS_LINE_2)
#             if phrase_removal != -1:
#                 ws_uasys['AS' + str(cell.row)].value = ADDRESS_LINE_1.strip(GOV_ADDRESS_LINE_2)
# # Move/delete substrings from CAMP_ADDRESS_LINE_2,
# for cell in ws_uasys['AU']:
#     CAMP_PO_BOX_LINE = str(cell.value).split()
#     word = CAMP_PO_BOX_LINE[0]
#     if word != 'PO' or word != 'N/A':
#         if word.find('STE') == -1:
#             try:
#                 ADDRESS_LINE_2 = ws_uasys['AT' + str(cell.row)].value
#                 ws_uasys['AU' + str(cell.row)].value = ADDRESS_LINE_2
#                 ws_uasys['AT' + str(cell.row)].value = 'N/A'
#                 ws_uasys['AT2'].value = 'CAMP_ADDRESS_LINE_2'
#             except AttributeError:
#                 print('MergedCell object attribute value is read-only')
#         else:
#             CAMP_POSTAL_CODE = ws_uasys['AV' + str(cell.row)].value
#             CAMP_MUNICIPALITY = ws_uasys['AU' + str(cell.row)].value
#             ws_uasys['AY' + str(cell.row)].value = CAMP_POSTAL_CODE
#             ws_uasys['AV' + str(cell.row)].value = CAMP_MUNICIPALITY
#             ws_uasys['AU' + str(cell.row)].value = 'N/A'
# for cell in ws_uasys['AY']:
#     POSTAL_CODE = str(cell.value)
#     try:
#         if POSTAL_CODE.isalpha() and POSTAL_CODE != 'N/A':
#             CAMP_MUNICIPALITY = ws_uasys['AY' + str(cell.row)].value
#             CAMP_ADDRESS_LINE_2 = ws_uasys['AV' + str(cell.row)].value
#             ws_uasys['AT' + str(cell.row)].value = CAMP_ADDRESS_LINE_2
#             ws_uasys['AV' + str(cell.row)].value = CAMP_MUNICIPALITY
#             ws_uasys['AY' + str(cell.row)].value = ''
#         if cell.value == 'N/A' or cell.value == 'N/':
#             CAMP_POSTAL_CODE = ws_uasys['AV' + str(cell.row)].value
#             ADDRESS_LINE_2 = ws_uasys['AU' + str(cell.row)].value
#             ws_uasys['AY' + str(cell.row)].value = CAMP_POSTAL_CODE
#             ws_uasys['AT' + str(cell.row)].value = ADDRESS_LINE_2
#             ws_uasys['AU' + str(cell.row)].value = 'N/A'
#             ws_uasys['AU2'].value = 'CAMP_PO_BOX_LINE'
#             ws_uasys['AV' + str(cell.row)].value = ''
#         postal_code_list = str(cell.value).split()
#         word = postal_code_list[0]
#         if word.isalpha() and len(word) <= 2:
#             STATE_REGION_SHORT = str(word).strip('[]')
#             ws_uasys['AW' + str(cell.row)].value = STATE_REGION_SHORT
#             ws_uasys['AY' + str(cell.row)].value = str(cell.value).strip(STATE_REGION_SHORT)
#     except IndexError:
#         print('list index out of range')
#     except AttributeError:
#         print('MergedCell object attribute value is read-only')
# # If CAMP_CAMPUS_NAME is the same as PRIMARY_INSTITUTION_NAME change camp name to Main Campus
# for cell in ws_uasys['AQ']:
#     campus_name = str(cell.value)
#     campus_name_list = str(cell.value).split()
#     institution_name = ws_uasys['U' + str(cell.row)].value
#     try:
#         for index in range(len(campus_name_list)):
#             word = campus_name_list[index]
#             first_at = word.find('at')
#             second_AT = word.find('AT')
#             if first_at == 0 or second_AT == 0:
#                 first_word = index + 1
#                 campus_name = str(' '.join(campus_name_list[first_word:len(campus_name_list)]))
#                 ws_uasys['AQ' + str(cell.row)].value = campus_name
#         if campus_name.upper() == institution_name.upper():
#             ws_uasys['AQ' + str(cell.row)].value = 'Main Campus'
#     except AttributeError:
#         print('NoneType object has no attribute upper')
# # Check to see if campus is inactive/closed according to NCES database
# for cell in ws_uasys['AP']:
#     organization_name = str(cell.value)
#     for look in ws_nces_grab['B']:
#         nces_institution = str(look.value)
#         if nces_institution.upper() == organization_name.upper():
#             institution_closed = ws_nces_grab['W' + str(look.row)].value
#             found_two = str(institution_closed).find('-2')
#             if found_two < 0:
#                 ws_uasys['BD' + str(cell.row)].value = institution_closed
#                 ws_uasys['BC' + str(cell.row)].value = 'Y'
# # If CAMPUS_RECORD_SOURCE is blank then assign the cell N/A
# for cell in ws_uasys['BE']:
#     try:
#         if cell.value is None:
#             ws_uasys['BE' + str(cell.row)].value = "N/A"
#     except AttributeError:
#         print('Cell is read only!')
#     except TypeError:
#         print('Cell is read only!')
#     except:
#         print('Unknown error')
# print('Done!')
wb_uasys.save(raw_file)
