from openpyxl import load_workbook
import openai

raw_file = input("File location in explorer(.xlsx): ")
wrong_input = raw_file.find(".xlsx")
if wrong_input == -1:
    print("Be sure to add .xlsx to end of file location!")
    raw_file = input("File location is explorer(.xlsx)")
wb_uasys = load_workbook(raw_file)
sheet_name = input("Name of sheet in raw file: ")
ws_uasys = wb_uasys[sheet_name]

for cell in ws_uasys['U3': ws_uasys.max_row]:
    institution_name = str(cell.value)
    municipality = str(ws_uasys['Y' + str(cell.row)].value)
    state = str(ws_uasys['Z' + str(cell.row)].value)

    API_KEY = open(r"C:\Users\Wayne\Work Stuff\Data Conversion\API Key.txt").read()
    openai.api_key = API_KEY
    response = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=[
            {"role": "user", "content": "Only return the date when " + institution_name + ' at ' + municipality + ',' + state + " was founded?"}
        ]
    )
    reply_content = response.choices[0].message.content
    print(reply_content)
wb_uasys.save(raw_file)
