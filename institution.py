from openpyxl import load_workbook
import re

raw_file = input("File location in explorer(.xlsx): ")
wrong_input = raw_file.find(".xlsx")
if wrong_input == -1:
    print("Be sure to add .xlsx to end of file location!")
    raw_file = input("File location is explorer(.xlsx)")
wb_uasys = load_workbook(raw_file)
wb_data_grab = load_workbook("AccreditationData.xlsx")
wb_nces_grab = load_workbook("Data_3-14-2023---623.xlsx")
sheet_name = input("Name of sheet in raw file: ")
ws_uasys = wb_uasys[sheet_name]
ws_data_grab = wb_data_grab["InstituteCampuses"]
ws_nces_grab = wb_nces_grab["Data_3-14-2023---623"]
# If PRIMARY_ORGANIZATION_ID is blank then assign the cell AutoGen
for cell in ws_uasys['Q']:
    try:
        if cell.value is None:
            ws_uasys['Q' + str(cell.row)].value = "AutoGen"
    except AttributeError:
        print(cell + ' is read only!')
    except TypeError:
        print('Cell is read only!')
    except:
        print('Unknown error')
# Get INST_PO_BOX_LINE for PRIMARY_INSTITUTION_NAME from LocationName -> Address
for cell in ws_uasys['U']:
    PRIMARY_INSTITUTION_NAME = str(cell.value).upper()
    cell_prev = int(cell.row) - 1
    try:
        if cell_prev != 0 and PRIMARY_INSTITUTION_NAME != ws_uasys['U' + str(cell_prev)].value.upper():
            print("----------------------------------")
            print(PRIMARY_INSTITUTION_NAME)
            for grab in ws_data_grab['D']:
                Location_Name = str(grab.value)
                if Location_Name.upper() == PRIMARY_INSTITUTION_NAME:
                    Address = str(ws_data_grab['H' + str(grab.row)].value)
                    if Address.find('P.O'):
                        found = re.search("x(.+?),", Address)
                        if not found:
                            continue
                        else:
                            number_POBOX = found.group(1)
                            INST_PO_BOX_LINE = str('PO Box' + str(number_POBOX))
                            print('Found: ' + INST_PO_BOX_LINE)
                            ws_uasys['X' + str(cell.row)].value = INST_PO_BOX_LINE
    except AttributeError:
        print("----------------------------------")
        print('NoneType for: ' + str(cell.value))
    except TypeError:
        print('NoneType')
    except:
        print('Unknown error')
# If INST_COUNTRY_CODE is blank then assign USA
for cell in ws_uasys['AA']:
    try:
        if cell.value is None:
            ws_uasys['AA' + str(cell.row)].value = "USA"
    except AttributeError:
        print('cell is read only!')
    except TypeError:
        print('Cell is read only!')
    except:
        print('Unknown error')
# Move/delete substrings from INST_ADDRESS_LINE_1 and moving them into respective column row
for cell in ws_uasys['V']:
    governing_address = str(cell.value).split()
    for index in range(len(governing_address)):
        word = governing_address[index]
        if word == 'Ste' or word == 'Ste.' or word == 'Unit' or word == 'PO' or word == 'Suite':
            INST_ADDRESS_LINE_1 = str(' '.join(governing_address[index:len(governing_address)]))
            found_pobox = INST_ADDRESS_LINE_1.find('PO Box')
            if found_pobox == -1:
                ws_uasys['W' + str(cell.row)].value = INST_ADDRESS_LINE_1.upper()
            else:
                ws_uasys['X' + str(cell.row)].value = INST_ADDRESS_LINE_1
                ws_uasys['W' + str(cell.row)].value = 'N/A'
            ADDRESS_LINE_1 = str(cell.value)
            phrase_removal = ADDRESS_LINE_1.find(INST_ADDRESS_LINE_1)
            if phrase_removal != -1:
                ws_uasys['V' + str(cell.row)].value = ADDRESS_LINE_1.strip(INST_ADDRESS_LINE_1)
        elif word == 'Floor' or word == 'Fl':
            floor_num = index - 1
            INST_ADDRESS_LINE_1 = str(' '.join(governing_address[floor_num:len(governing_address)]))
            ws_uasys['W' + str(cell.row)].value = INST_ADDRESS_LINE_1.upper()
            ADDRESS_LINE_1 = str(cell.value)
            phrase_removal = ADDRESS_LINE_1.find(INST_ADDRESS_LINE_1)
            if phrase_removal != -1:
                ws_uasys['V' + str(cell.row)].value = ADDRESS_LINE_1.strip(INST_ADDRESS_LINE_1)
# If INST_ADDRESS_LINE_2 is blank then assign the cell N/A
for cell in ws_uasys['W']:
    try:
        if cell.value is None:
            ws_uasys['W' + str(cell.row)].value = "N/A"
    except AttributeError:
        print('Cell is read only!')
    except TypeError:
        print('Cell is read only!')
    except:
        print('Unknown error')
# Check to see if institution is inactive/closed according to NCES database
for cell in ws_uasys['U']:
    organization_name = str(cell.value)
    cell_prev = int(cell.row) - 1
    try:
        if cell_prev != 0 and organization_name != ws_uasys['U' + str(cell_prev)].value.upper():
            for look in ws_nces_grab['B']:
                nces_institution = str(look.value)
                if nces_institution.upper() == organization_name.upper():
                    institution_closed = ws_nces_grab['W' + str(look.row)].value
                    found_two = str(institution_closed).find('-2')
                    if found_two < 0:
                        ws_uasys['AI' + str(cell.row)].value = institution_closed
    except AttributeError:
        print("----------------------------------")
        print('NoneType for: ' + str(cell.value))
    except TypeError:
        print('NoneType')
    except:
        print('Unknown error')
# if INST_RECORD_SOURCE is blank then assign N/A
for cell in ws_uasys['AJ']:
    try:
        if cell.value is None:
            ws_uasys['AJ' + str(cell.row)].value = "N/A"
    except AttributeError:
        print('cell is read only!')
    except TypeError:
        print('Cell is read only!')
    except:
        print('Unknown error')
print('Done!')
wb_uasys.save(raw_file)
