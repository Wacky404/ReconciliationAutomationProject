from openpyxl import load_workbook
import automation

wb_uasys = load_workbook(r"C:\Users\Wayne Cole\Downloads\Work Stuff\Copy TexasEducationalInstitutionsDatabase.xlsx")
wb_data_grab = load_workbook(r"C:\Users\Wayne Cole\Downloads\Work Stuff\AccreditationData.xlsx")
wb_nces_grab = load_workbook(r"C:\Users\Wayne Cole\Downloads\Work Stuff\Data_3-14-2023---623")
ws_uasys = wb_uasys["All Texas Institutions"]
ws_data_grab = wb_data_grab["InstituteCampuses"]
ws_nces_grab = wb_nces_grab["in"]

# governing institution automation
automation.governing_id(ws_uasys)
automation.governing_name(ws_uasys, ws_data_grab)
print("Populating associated fields.....hold on.....")
automation.governing_edids(ws_uasys, ws_data_grab)
automation.governing_location(ws_uasys, ws_data_grab)
automation.governing_region_short(ws_uasys)
automation.governing_phone(ws_uasys, ws_data_grab)
automation.governing_recordsource(ws_uasys)

# institution automation
automation.institution_id(ws_uasys)
automation.institution_po(ws_uasys, ws_data_grab)
automation.institution_country_code(ws_uasys)
automation.institution_established_date(ws_uasys)
automation.institution_recordsource(ws_uasys)

# campus institution automation
automation.campus_id(ws_uasys)
automation.campus_institution_edids(ws_uasys, ws_data_grab)
automation.campus_po_phonenumber(ws_uasys, ws_data_grab)
automation.campus_recordsource(ws_uasys)

print('Done!')
wb_uasys.save(r"C:\Users\Wayne Cole\Downloads\Work Stuff\Copy TexasEducationalInstitutionsDatabase.xlsx")
