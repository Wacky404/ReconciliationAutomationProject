from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from difflib import SequenceMatcher
from src.utils.log_util import logger
from src.NominatimIntegration import NominatimIntegration
import src.place_id as place_id
import os.path as osp
import re


class DataFile:
    wb_data_grab = load_workbook('data/AccreditationData.xlsx')
    wb_nces_grab = load_workbook('data/Data_3-14-2023---623.xlsx')
    ws_data_grab = wb_data_grab["InstituteCampuses"]
    ws_nces_grab = wb_nces_grab["Data_3-14-2023---623"]
    full_spellings: dict = {
        'rd': 'road',
        'rd.': 'road',
        'ave': 'avenue',
        'ave.': 'avenue',
        'dr': 'drive',
        'dr.': 'drive',
        'st': 'street',
        'st.': 'street',
        'str': 'street',
        'hwy': 'highway',
        'hwy.': 'highway',
        'blvd': 'boulevard',
        'blvd.': 'boulevard',
        'tr': 'trail',
        'tr.': 'trail',
        'n': 'north',
        'n.': 'north',
        'e': 'east',
        'e.': 'east',
        's': 'south',
        's.': 'south',
        'w': 'west',
        'w.': 'west',
        'sw': 'southwest',
        's.w.': 'southwest',
        's.w': 'southwest',
        'se': 'southeast',
        's.e.': 'southeast',
        's.e': 'southeast',
        'nw': 'northwest',
        'n.w.': 'northwest',
        'n.w': 'northwest',
        'ne': 'northeast',
        'n.e.': 'northeast',
        'n.e': 'northeast',
        'pky': 'parkway',
        'pky.': 'parkway',
        'sr': 'state highway system',
        'sr.': 'state highway system',
        'us': 'united states',
        'u.s.': 'united states',
        'u.s': 'united states'
    }

    gov_field_names: dict = {
        'F': 'gov_address_line_1',
        'I': 'gov_municipality',
        'L': 'gov_postal_code',
    }
    insti_field_names: dict = {
        'U': 'primary_institution_name',
        'V': 'inst_address_line_1',
        'Y': 'inst_municipality',
        'AB': 'inst_postal_code',
    }
    camp_field_names: dict = {
        'AP': 'camp_official_institution_name',
        'AQ': 'camp_campus_name',
        'AR': 'camp_location_site',
        'AS': 'camp_address_line_1',
        'AV': 'camp_municipality',
        'AY': 'camp_postal_code',
    }
    null_values: tuple = ('NONE', 'None', 'NULL', 'Null', '')

    def __init__(self, raw_file, sheet_name, abbrev):
        self.raw_file = raw_file
        # Check on this - Wayne
        base_file: str = osp.basename(self.raw_file)
        transf_f: str = osp.join(osp.expanduser(
            '~'), 'Documents', 'PipelineOutput', base_file)
        self.transf_file = transf_f if transf_f is not None else self.raw_file
        self.sheet_name = sheet_name
        self.abbrev = abbrev
        self.wb_uasys = load_workbook(raw_file)
        self.ws_uasys = self.wb_uasys[sheet_name]

    @staticmethod
    def raw_file_check(raw_file: str) -> str:
        wrong_input = raw_file.find(".xlsx")
        if wrong_input == -1:
            print("Be sure to add .xlsx to end of file location!")
            raw_file = input("File location is explorer(.xlsx)")
            return raw_file

    @staticmethod
    def has_numbers(input_string: str) -> bool:
        return any(char.isdigit() for char in input_string)

    @classmethod
    def _split_address(cls, ws_uasys, address_original, cache_row, addr_line1_col, addr_line2_col) -> None:
        split_address_words = (
            'Ste',
            'Ste.',
            'STE',
            'STE.',
            'Unit',
            'PO',
            'Po',
            'Suite',
            'suite',
        )
        secondary_split_address_words = ('Floor', 'Fl', 'fl')
        address = address_original
        for index in range(len(address)):
            word = address[index]

            for match in split_address_words:
                if word == match:
                    address_line_2 = str(' '.join(address[index:len(address)]))
                    ws_uasys[str(addr_line2_col) + str(cache_row)
                             ].value = address_line_2
                    address_line_1 = str(address)
                    phrase_removal = address_line_1.find(address_line_2)
                    if phrase_removal != -1:
                        ws_uasys[str(addr_line1_col) + str(cache_row)
                                 ].value = address_line_1.strip(address_line_2)

            for other in secondary_split_address_words:
                if word == other:
                    floor_num = index - 1
                    address_line_2 = str(
                        ' '.join(address[floor_num:len(address)]))
                    ws_uasys[str(addr_line2_col) + str(cache_row)
                             ].value = address_line_2.upper()
                    address_line_1 = str(address)
                    phrase_removal = address_line_1.find(address_line_2)
                    if phrase_removal != -1:
                        ws_uasys[str(addr_line1_col) + str(cache_row)
                                 ].value = address_line_1.strip(address_line_2)

    @classmethod
    def reconcile_institution(cls, wb_uasys, ws_uasys, file_location, ws_data_grab, ws_nces_grab) -> None:
        # Inputs Autogen in field cells
        logger.info('Inputting Autogen in field cells...')
        for cell in ws_uasys['Q']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['Q' + str(cell.row)].value = "AutoGen"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        # Get inst_po_box_line for primary_institution_name from LocationName -> address
        logger.info('Getting Institution PO_BOX for primary institution...')
        for cell in ws_uasys['U']:
            if cell.row >= 3:
                primary_institution_name = str(cell.value).upper()
                cell_prev = int(cell.row) - 1
                try:
                    if cell_prev != 0 and primary_institution_name != ws_uasys['U' + str(cell_prev)].value.upper():
                        logger.debug(primary_institution_name)

                        for grab in ws_data_grab['D']:
                            location_name = str(grab.value)
                            if location_name.upper() == primary_institution_name:
                                address = str(
                                    ws_data_grab['H' + str(grab.row)].value)
                                if address.find('P.O'):
                                    found = re.search("x(.+?),", address)
                                    if not found:
                                        continue
                                    else:
                                        number_pobox = found.group(1)
                                        inst_po_box_line = str(
                                            'PO Box' + str(number_pobox))
                                        logger.debug(
                                            'Found: ' + inst_po_box_line)
                                        ws_uasys['X' + str(cell.row)
                                                 ].value = inst_po_box_line
                except Exception as e:
                    logger.exception(
                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        # If INST_COUNTRY_CODE is blank then assign USA
        logger.info('If Institution COUNTRY_CODE is blank then assign USA...')
        for cell in ws_uasys['AA']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['AA' + str(cell.row)].value = "USA"
                except Exception as e:
                    logger.exception(
                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        # Move/delete substrings from inst_address_line_1 and moving them into respective column row
        logger.info('Moving/Deleting substrings from Institution address line...')
        substrings = {
            'Ste',
            'Ste.',
            'STE',
            'STE.',
            'Unit',
            'PO',
            'Po',
            'Suite',
            'suite',
            'Building',
        }
        for cell in ws_uasys['V']:
            if cell.row >= 3:
                governing_address = str(cell.value).split()

                for index in range(len(governing_address)):
                    word = governing_address[index]

                    for look in substrings:
                        if word == look:
                            inst_address_line_1 = str(
                                ' '.join(governing_address[index:len(governing_address)]))
                            found_pobox = inst_address_line_1.find('PO Box')
                            if found_pobox == -1:
                                ws_uasys['W' + str(cell.row)
                                         ].value = inst_address_line_1.upper()
                            else:
                                ws_uasys['X' + str(cell.row)
                                         ].value = inst_address_line_1
                                ws_uasys['W' + str(cell.row)].value = 'N/A'
                            address_line_1 = str(cell.value)
                            phrase_removal = address_line_1.find(
                                inst_address_line_1)
                            if phrase_removal != -1:
                                ws_uasys['V' + str(cell.row)].value = address_line_1.strip(
                                    inst_address_line_1)
                        elif word == 'Floor' or word == 'Fl':
                            floor_num = index - 1
                            inst_address_line_1 = str(
                                ' '.join(governing_address[floor_num:len(governing_address)]))
                            ws_uasys['W' + str(cell.row)
                                     ].value = inst_address_line_1.upper()
                            address_line_1 = str(cell.value)
                            phrase_removal = address_line_1.find(
                                inst_address_line_1)
                            if phrase_removal != -1:
                                ws_uasys['V' + str(cell.row)].value = address_line_1.strip(
                                    inst_address_line_1)
                            break
        # If INST_ADDRESS_LINE_2 is blank then assign the cell N/A
        logger.info('If Institution ADDRESS_LINE_2 is blank then assigning cell N/A...')
        for cell in ws_uasys['W']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['W' + str(cell.row)].value = "N/A"
                except Exception as e:
                    logger.exception(
                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        # Check to see if institution is inactive/closed according to NCES database
        logger.info('Checking to see if Institution is inactive/closed according to NCES database...')
        for cell in ws_uasys['U']:
            if cell.row >= 3:
                organization_name = str(cell.value)
                cell_prev = int(cell.row) - 1
                try:
                    if cell_prev != 0 and organization_name != ws_uasys['U' + str(cell_prev)].value.upper():

                        for look in ws_nces_grab['B']:
                            nces_institution = str(look.value)
                            if nces_institution.upper() == organization_name.upper():
                                institution_closed = ws_nces_grab['W' + str(
                                    look.row)].value
                                found_two = str(institution_closed).find('-2')
                                if found_two < 0:
                                    ws_uasys['AI' + str(cell.row)
                                             ].value = institution_closed
                                    ws_uasys['AH' + str(cell.row)].value = 'Y'
                except Exception as e:
                    logger.exception(
                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        # if INST_RECORD_SOURCE is blank then assign N/A
        logger.info('If Institution RECORD_SOURCE is blank then assigning cell(s) N/A...')
        for cell in ws_uasys['AJ']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['AJ' + str(cell.row)].value = "N/A"
                except Exception as e:
                    logger.exception(
                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        wb_uasys.save(file_location)

    @classmethod
    def clean_institution(cls, wb_uasys, ws_uasys, file_location, full_spellings):
        for cell in ws_uasys['R']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['R' + str(cell.row)].value = "NULL"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['S']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['S' + str(cell.row)].value = "NULL"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['T']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['T' + str(cell.row)].value = "NULL"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AC']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['AC' + str(cell.row)].value = "NULL"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        yellow = 'FFFF00'
        red = 'FF6666'
        y_highlight = PatternFill(patternType='solid', fgColor=yellow)
        r_highlight = PatternFill(patternType='solid', fgColor=red)
        for cell in ws_uasys['Q']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['Q' + str(cell.row)].value = 'AutoGen'
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['R']:
            try:
                if cell.row >= 3:
                    gov_id = str(cell.value)
                    gov_id_oped = str(ws_uasys['S' + str(cell.row)].value)
                    gov_id_iped = str(ws_uasys['T' + str(cell.row)].value)
                    if not gov_id.isnumeric():
                        ws_uasys['R' + str(cell.row)].fill = r_highlight
                    if not gov_id_oped.isnumeric():
                        ws_uasys['S' + str(cell.row)].fill = r_highlight
                    if not gov_id_iped.isnumeric():
                        ws_uasys['T' + str(cell.row)].fill = r_highlight
            except:
                logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['U']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['U' + str(cell.row)].fill = y_highlight
                    else:
                        ws_uasys['U' + str(cell.row)
                                 ].value = str(cell.value).upper()
                except:
                    logger.exception(f'Error with {cell.coordinaate}')
        for cell in ws_uasys['V']:
            if cell.row >= 3:
                try:
                    ws_uasys['V' + str(cell.row)
                             ].value = str(cell.value).upper()
                    if cell.value is None:
                        ws_uasys['V' + str(cell.row)].fill = r_highlight
                    gov_address = str(
                        ws_uasys['V' + str(cell.row)].value).lower()
                    sep_address = gov_address.split()

                    for key in full_spellings:

                        for index in range(len(sep_address)):
                            word = sep_address[index]
                            if word == key:
                                sep_address[index] = full_spellings[key]
                                gov_address = str(' '.join(sep_address))
                                ws_uasys['V' + str(cell.row)
                                         ].value = gov_address.upper()
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['W']:
            if cell.row >= 3:
                try:
                    address_one = cell.value
                    if address_one is None:
                        ws_uasys['W' + str(cell.row)].value = 'N/A'
                    elif address_one.find('PO') != -1:
                        ws_uasys['W' + str(cell.row)].fill = y_highlight
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['X']:
            if cell.row >= 3:
                try:
                    address_two = str(cell.value)
                    if cell.value is None:
                        ws_uasys['X' + str(cell.row)].value = 'N/A'
                        wb_uasys.save(file_location)
                    elif address_two.find('PO') == -1:
                        ws_uasys['X' + str(cell.row)].fill = y_highlight
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['Y']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['Y' + str(cell.row)].fill = r_highlight
                    else:
                        ws_uasys['Y' + str(cell.row)
                                 ].value = str(cell.value).upper()
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['Z']:
            try:
                if cell.row >= 3:
                    region = str(cell.value)
                    if len(region) != 2:
                        ws_uasys['Z' + str(cell.row)].fill = r_highlight
            except:
                logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AA']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['AA' + str(cell.row)].value = 'USA'
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AB']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['AB' + str(cell.row)].fill = r_highlight
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AF']:
            if cell.row >= 3:
                try:
                    if cell.value == 'Manually Find' or cell.value is None:
                        ws_uasys['AF' + str(cell.row)].fill = r_highlight
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AG']:
            if cell.row >= 3:
                try:
                    if cell.value == 'Manually Find' or cell.value is None:
                        ws_uasys['AG' + str(cell.row)].fill = r_highlight
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AH']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['AH' + str(cell.row)].value = "N"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AI']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['AI' + str(cell.row)].value = "N/A"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AJ']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['AJ' + str(cell.row)].value = "N/A"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        wb_uasys.save(file_location)

    @classmethod
    def reconcile_governing(cls, wb_uasys, ws_uasys, file_location, abbrev, ws_data_grab, ws_nces_grab):
        # If GOVERNING_ORGANIZATION_ID is blank then assign the cell AutoGen
        logger.info('If Governing ORGANIZATION_ID is blank then assign the cell AutoGen..')
        for cell in ws_uasys['A']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['A' + str(cell.row)].value = "AutoGen"
                except Exception as e:
                    logger.exception(
                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        # Get primary institution name and compare it against cells in additional sites location name,
        # if match: access Parent
        # Name Cell and return cell data to populate Governing Org name of same row as primary institution name
        logger.info('Getting primary Institution name and comparing it against cell in additional sites locations...')
        for cell in ws_uasys['U']:
            if cell.row >= 3:
                institute_name = str(cell.value)
                logger.debug(institute_name)

                for grab in ws_data_grab['D']:
                    location_name = str(grab.value)
                    if location_name.upper() == institute_name.upper():
                        parent_name = str(
                            ws_data_grab['E' + str(grab.row)].value)
                        if parent_name == "-":
                            ws_uasys['E' + str(cell.row)
                                     ].value = institute_name
                        else:
                            ws_uasys['E' + str(cell.row)].value = parent_name
                        logger.debug('Placed in governing: ' +
                                     str(ws_uasys['E' + str(cell.row)].value))
        # Get Governing_Organization_Name's DAPIP, OPE, and IPEDSID IDs from data_grab
        logger.info('Getting Governing ORGANIZATION_NAMEs DAPID, OPE, and IPEDSID ids from data source...')
        for cell in ws_uasys['E']:
            if cell.row >= 3:
                institution_govern = str(cell.value)

                for grab in ws_data_grab['E']:
                    location_name = str(grab.value)
                    if location_name.upper() == institution_govern.upper():
                        GOV_DAPID = str(
                            ws_data_grab['F' + str(grab.row)].value)
                        GOV_DAPID = GOV_DAPID if GOV_DAPID != '-' else 'NULL'
                        ws_uasys['B' + str(cell.row)].value = GOV_DAPID
                        # Checking for the rest of the IDs in the accred database
                        if GOV_DAPID == 'NULL':
                            govern_zipcode = str(
                                ws_uasys['L' + str(cell.row)].value)

                            for look in ws_nces_grab['B']:
                                match_institution = str(look.value)
                                nces_zipcode = str(
                                    ws_nces_grab['K' + str(look.row)].value)
                                if match_institution.upper() == institution_govern.upper() and govern_zipcode == nces_zipcode:
                                    found_zipcode = str(
                                        ws_nces_grab['A' + str(look.row)].value)
                                    ws_uasys['B' + str(cell.row)
                                             ].value = found_zipcode
                        else:
                            for match in ws_data_grab['A']:
                                accred_dapid = str(match.value)
                                if accred_dapid == GOV_DAPID:
                                    GOV_OPEID = str(
                                        ws_data_grab['B' + str(match.row)].value)
                                    GOV_IPEDID = str(
                                        ws_data_grab['C' + str(match.row)].value)
                                    ws_uasys['C' +
                                             str(cell.row)].value = GOV_OPEID
                                    ws_uasys['D' +
                                             str(cell.row)].value = GOV_IPEDID
        # Get GOV address line 1, GOV_MUNICIPALITY, GOV postal code
        logger.info('Getting Governing ADDRESS_LINE_1, Governing MUNICIPALITY, Governing POSTAL_CODE...')
        for cell in ws_uasys['B']:
            if cell.row >= 3:
                institution_govern = str(cell.value)

                for grab in ws_data_grab['A']:
                    location_name = str(grab.value)
                    if location_name == institution_govern:
                        address_grab = str(
                            ws_data_grab['H' + str(grab.row)].value)
                        address_lst = address_grab.split(', ', maxsplit=7)
                        try:
                            if len(address_lst) == 1:
                                address_grab = address_grab + ", NULL, NULL, NULL, NULL, NULL, NULL"
                            elif len(address_lst) == 2:
                                address_grab = address_grab + ", NULL, NULL, NULL, NULL, NULL"
                            elif len(address_lst) == 3:
                                address_grab = address_grab + ", NULL, NULL, NULL, NULL"
                            elif len(address_lst) == 4:
                                address_grab = address_grab + ", NULL, NULL, NULL"

                            if len(address_grab.split(', ')) == 7:
                                GOV_ADDRESS_LINE_1, temp_LINE_2, temp_POBOX, temp_MUNI, temp_PCODE, temp1_Unknown, \
                                    temp2_Unknown = address_grab.split(', ')

                                if GOV_ADDRESS_LINE_1.startswith('P.O. Box'):
                                    temp_PCODE = temp_POBOX
                                    temp_POBOX = GOV_ADDRESS_LINE_1
                                    GOV_ADDRESS_LINE_1 = 'N/A'

                                if temp_POBOX.startswith('K'):
                                    temp_POBOX = 'N/A'
                                    temp_MUNI = temp_PCODE
                                    temp_PCODE = temp1_Unknown

                                if temp_PCODE.startswith('P.O BOX'):
                                    temp_POBOX = temp_PCODE
                                    temp_PCODE = 'NULL'

                                if not temp_LINE_2.startswith('Suite'):
                                    temp_MUNI = temp_LINE_2
                                    temp_LINE_2 = 'N/A'
                                    temp_PCODE = temp_POBOX
                                    temp_POBOX = 'N/A'
                                if temp_MUNI.startswith(abbrev.upper()):
                                    temp_PCODE = temp_MUNI
                                    temp_MUNI = temp_POBOX
                                    temp_POBOX = 'N/A'

                                GOV_ADDRESS_LINE_2 = temp_LINE_2.upper()
                                GOV_PO_BOX_LINE = temp_POBOX.strip('.')
                                GOV_MUNICIPALITY = temp_MUNI.upper()
                                GOV_POSTAL_CODE = temp_PCODE.strip(
                                    abbrev.upper())

                                ws_uasys['F' + str(cell.row)
                                         ].value = GOV_ADDRESS_LINE_1
                                ws_uasys['G' + str(cell.row)
                                         ].value = GOV_ADDRESS_LINE_2
                                ws_uasys['H' + str(cell.row)
                                         ].value = GOV_PO_BOX_LINE
                                ws_uasys['I' + str(cell.row)
                                         ].value = GOV_MUNICIPALITY
                                ws_uasys['L' + str(cell.row)
                                         ].value = GOV_POSTAL_CODE

                            elif len(address_grab.split(', ')) == 6:
                                GOV_ADDRESS_LINE_1, temp_LINE_2, temp_POBOX, temp_MUNI, temp_PCODE, temp1_Unknown = \
                                    address_grab.split(', ')

                                if GOV_ADDRESS_LINE_1.startswith('P.O. Box'):
                                    temp_PCODE = temp_POBOX
                                    temp_POBOX = GOV_ADDRESS_LINE_1
                                    GOV_ADDRESS_LINE_1 = 'N/A'

                                if temp_POBOX.startswith('K'):
                                    temp_POBOX = 'N/A'
                                    temp_MUNI = temp_PCODE
                                    temp_PCODE = temp1_Unknown

                                if temp_PCODE.startswith('P.O BOX'):
                                    temp_POBOX = temp_PCODE
                                    temp_PCODE = 'NULL'

                                if not temp_LINE_2.startswith('Suite'):
                                    temp_MUNI = temp_LINE_2
                                    temp_LINE_2 = 'N/A'
                                    temp_PCODE = temp_POBOX
                                    temp_POBOX = 'N/A'
                                if temp_MUNI.startswith(abbrev.upper()):
                                    temp_PCODE = temp_MUNI
                                    temp_MUNI = temp_POBOX
                                    temp_POBOX = 'N/A'

                                GOV_ADDRESS_LINE_2 = temp_LINE_2.upper()
                                GOV_PO_BOX_LINE = temp_POBOX.strip('.')
                                GOV_MUNICIPALITY = temp_MUNI.upper()
                                GOV_POSTAL_CODE = temp_PCODE.strip(
                                    abbrev.upper())

                                ws_uasys['F' + str(cell.row)
                                         ].value = GOV_ADDRESS_LINE_1
                                ws_uasys['G' + str(cell.row)
                                         ].value = GOV_ADDRESS_LINE_2
                                ws_uasys['H' + str(cell.row)
                                         ].value = GOV_PO_BOX_LINE
                                ws_uasys['I' + str(cell.row)
                                         ].value = GOV_MUNICIPALITY
                                ws_uasys['L' + str(cell.row)
                                         ].value = GOV_POSTAL_CODE

                            elif len(address_grab.split(', ')) == 5:
                                GOV_ADDRESS_LINE_1, temp_LINE_2, temp_POBOX, temp_MUNI, temp_PCODE = \
                                    address_grab.split(', ')

                                if GOV_ADDRESS_LINE_1.startswith('P.O. Box'):
                                    temp_PCODE = temp_POBOX
                                    temp_POBOX = GOV_ADDRESS_LINE_1
                                    GOV_ADDRESS_LINE_1 = 'N/A'

                                if temp_PCODE.startswith('P.O BOX'):
                                    temp_POBOX = temp_PCODE
                                    temp_PCODE = 'NULL'

                                if not temp_LINE_2.startswith('Suite'):
                                    temp_MUNI = temp_LINE_2
                                    temp_LINE_2 = 'N/A'
                                    temp_PCODE = temp_POBOX
                                    temp_POBOX = 'N/A'
                                if temp_MUNI.startswith(abbrev.upper()):
                                    temp_PCODE = temp_MUNI
                                    temp_MUNI = temp_POBOX
                                    temp_POBOX = 'N/A'

                                GOV_ADDRESS_LINE_2 = temp_LINE_2.upper()
                                GOV_PO_BOX_LINE = temp_POBOX.strip('.')
                                GOV_MUNICIPALITY = temp_MUNI.upper()
                                GOV_POSTAL_CODE = temp_PCODE.strip(
                                    abbrev.upper())

                                ws_uasys['F' + str(cell.row)
                                         ].value = GOV_ADDRESS_LINE_1
                                ws_uasys['G' + str(cell.row)
                                         ].value = GOV_ADDRESS_LINE_2
                                ws_uasys['H' + str(cell.row)
                                         ].value = GOV_PO_BOX_LINE
                                ws_uasys['I' + str(cell.row)
                                         ].value = GOV_MUNICIPALITY
                                ws_uasys['L' + str(cell.row)
                                         ].value = GOV_POSTAL_CODE

                        except Exception as e:
                            logger.exception(
                                f"An exception of type {type(e).__name__} occurred, NULL assigned. Details: {str(e)}")
                            ws_uasys['F' + str(cell.row)].value = 'NULL'
                            ws_uasys['G' + str(cell.row)].value = 'NULL'
                            ws_uasys['H' + str(cell.row)].value = 'NULL'
                            ws_uasys['I' + str(cell.row)].value = 'NULL'
                            ws_uasys['L' + str(cell.row)].value = 'NULL'
        # If GOV_STATE_REGION_SHORT is blank then assign worksheet state
        logger.info('If Governing STATE_REGION_SHORT is blank then assign worksheet state...')
        for cell in ws_uasys['J']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['J' + str(cell.row)].value = abbrev.upper()
                except Exception as e:
                    logger.exception(
                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        # If GOV_COUNTRY_CODE is blank then assign USA
        logger.info('If Governing COUNTRY_CODE is blank then assign USA...')
        for cell in ws_uasys['K']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['K' + str(cell.row)].value = "USA"
                except Exception as e:
                    logger.exception(
                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        # Get GOV_PhoneNumberFull
        logger.info('Getting Governing PhoneNumberFull...')
        for cell in ws_uasys['E']:
            if cell.row >= 3:
                institution_govern = str(cell.value)

                for grab in ws_data_grab['D']:
                    location_name = str(grab.value)
                    if location_name.upper() == institution_govern.upper():
                        phoneNumber_grab = str(
                            ws_data_grab['I' + str(grab.row)].value)
                        ws_uasys['M' + str(cell.row)].value = phoneNumber_grab
                        if phoneNumber_grab is None:
                            logger.debug(
                                'No phone number from Accreditation Database : Searching')

                            for look in ws_nces_grab['B']:
                                nces_institution = str(look.value)
                                if nces_institution.upper() == institution_govern.upper():
                                    phoneNumber_grab = str(
                                        ws_nces_grab['L' + str(grab.row)].value)
                                    ws_uasys['M' + str(cell.row)
                                             ].value = phoneNumber_grab
                                    logger.debug(
                                        'Found phone number in NCES Database')
                                else:
                                    logger.debug('No phone number found')
        # Check to see if GOV_ORG is inactive/closed according to NCES database
        logger.info('Checking to see if Governing ORGANIZATION is inactive/closed according to NCES database...')
        for cell in ws_uasys['E']:
            if cell.row >= 3:
                institution_govern = str(cell.value)

                for look in ws_nces_grab['B']:
                    nces_institution = str(look.value)
                    if nces_institution.upper() == institution_govern.upper():
                        institution_closed = ws_nces_grab['W' +
                                                          str(look.row)].value
                        found_two = str(institution_closed).find('-2')
                        if found_two < 0:
                            ws_uasys['O' + str(cell.row)
                                     ].value = institution_closed
        # If GOV_RECORD_SOURCE is blank then assign the cell N/A
        logger.info('If Governing RECORD_SOURCE is blank then assign the cell N/A...')
        for cell in ws_uasys['P']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['P' + str(cell.row)].value = "N/A"
                except Exception as e:
                    logger.exception(
                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        # if not in data_grab then search nces_grab database
        logger.info('If not in data source then search nces database...')
        for cell in ws_uasys['E']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        search_institution = ws_uasys['U' +
                                                      str(cell.row)].value

                        for look in ws_nces_grab['B']:
                            nces_institution = str(look.value)
                            if nces_institution.upper() == search_institution.upper():
                                GOV_DAPID = ws_uasys['R' + str(cell.row)].value
                                GOV_OPEID = ws_uasys['S' + str(cell.row)].value
                                GOV_IPEDID = ws_uasys['T' +
                                                      str(cell.row)].value
                                PRIMARY_INSTITUTION_NAME = nces_institution.upper()
                                GOV_ADDRESS_LINE_1 = ws_nces_grab['I' + str(
                                    look.row)].value
                                GOV_MUNICIPALITY = ws_nces_grab['J' +
                                                                str(look.row)].value
                                GOV_STATE_REGION_SHORT = ws_nces_grab['C' + str(
                                    look.row)].value
                                GOV_POSTAL_CODE = ws_nces_grab['K' +
                                                               str(look.row)].value
                                GOV_PhoneNumberFull = ws_nces_grab['L' + str(
                                    look.row)].value

                                ws_uasys['B' + str(cell.row)].value = GOV_DAPID
                                ws_uasys['C' + str(cell.row)].value = GOV_OPEID
                                ws_uasys['D' + str(cell.row)
                                         ].value = GOV_IPEDID
                                ws_uasys['E' + str(cell.row)
                                         ].value = PRIMARY_INSTITUTION_NAME.upper()
                                ws_uasys['F' + str(cell.row)
                                         ].value = GOV_ADDRESS_LINE_1
                                ws_uasys['I' + str(cell.row)
                                         ].value = GOV_MUNICIPALITY
                                ws_uasys['J' + str(cell.row)
                                         ].value = GOV_STATE_REGION_SHORT
                                ws_uasys['L' + str(cell.row)
                                         ].value = GOV_POSTAL_CODE
                                ws_uasys['M' + str(cell.row)
                                         ].value = GOV_PhoneNumberFull
                except Exception as e:
                    logger.exception(
                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        # Move/delete substrings from GOV_ADDRESS_LINE_1 and moving them into respective column row
        logger.info('Moving/deleting substrings from Governing ADDRESS_LINE_1...')
        substrings = {
            'Ste',
            'Ste.',
            'STE',
            'STE.',
            'Unit',
            'PO',
            'Po',
            'Suite',
            'suite'
        }
        for cell in ws_uasys['F']:
            if cell.row >= 3:
                governing_address = str(cell.value).split()

                for index in range(len(governing_address)):
                    word = governing_address[index]

                    for look in substrings:
                        if word == look:
                            GOV_ADDRESS_LINE_2 = str(
                                ' '.join(governing_address[index:len(governing_address)]))
                            found_pobox = GOV_ADDRESS_LINE_2.find('PO Box')
                            if found_pobox == -1:
                                ws_uasys['G' + str(cell.row)
                                         ].value = GOV_ADDRESS_LINE_2.upper()
                            else:
                                ws_uasys['H' + str(cell.row)
                                         ].value = GOV_ADDRESS_LINE_2
                                ws_uasys['G' + str(cell.row)].value = 'N/A'
                            ADDRESS_LINE_1 = str(cell.value)
                            phrase_removal = ADDRESS_LINE_1.find(
                                GOV_ADDRESS_LINE_2)
                            if phrase_removal != -1:
                                ws_uasys['F' + str(cell.row)].value = ADDRESS_LINE_1.strip(
                                    GOV_ADDRESS_LINE_2)
                        elif word == 'Floor' or word == 'Fl':
                            floor_num = index - 1
                            GOV_ADDRESS_LINE_2 = str(
                                ' '.join(governing_address[floor_num:len(governing_address)]))
                            ws_uasys['G' + str(cell.row)
                                     ].value = GOV_ADDRESS_LINE_2.upper()
                            ADDRESS_LINE_1 = str(cell.value)
                            phrase_removal = ADDRESS_LINE_1.find(
                                GOV_ADDRESS_LINE_2)
                            if phrase_removal != -1:
                                ws_uasys['F' + str(cell.row)].value = ADDRESS_LINE_1.strip(
                                    GOV_ADDRESS_LINE_2)
        # Move/delete substrings from GOV_POSTAL_CODE to GOV_MUNICIPALITY, GOV_MUNICIPALITY moves to GOV_ADDRESS_LINE_2
        logger.info('Moving/deleting substrings from Governing POSTAL_CODE and Governing MUNICIPALITY...')
        for cell in ws_uasys['L']:
            if cell.row >= 3:
                postal_code = str(cell.value).split()

                for index in range(len(postal_code)):
                    word = postal_code[index]
                    if not word.isalpha():
                        if word == "N/":
                            gov_municipality = ws_uasys['I' +
                                                        str(cell.row)].value.split()
                            ws_uasys['J' + str(cell.row)
                                     ].value = str(gov_municipality[0])
                            ws_uasys['L' + str(cell.row)
                                     ].value = str(gov_municipality[1])
                            GOV_MUNICIPALITY = ws_uasys['H' +
                                                        str(cell.row)].value
                            ws_uasys['I' + str(cell.row)
                                     ].value = GOV_MUNICIPALITY
                            ws_uasys['H' + str(cell.row)].value = 'N/A'
                    else:
                        GOV_POSTAL_CODE = str(
                            ' '.join(postal_code[index:len(postal_code)]))
                        if GOV_POSTAL_CODE.isalpha():
                            try:
                                GOV_ADDRESS_LINE_1 = ws_uasys['I' +
                                                              str(cell.row)].value
                                GOV_MUNICIPALITY = ws_uasys['L' +
                                                            str(cell.row)].value
                                ws_uasys['I' + str(cell.row)
                                         ].value = GOV_MUNICIPALITY
                                ws_uasys['G' + str(cell.row)
                                         ].value = GOV_ADDRESS_LINE_1
                                ws_uasys['L' + str(cell.row)].value = ''
                            except Exception as e:
                                logger.exception(
                                    f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
                        else:
                            GOV_STATE_REGION_SHORT = str(postal_code[index])
                            ws_uasys['J' + str(cell.row)
                                     ].value = GOV_STATE_REGION_SHORT
                            ws_uasys['L' + str(cell.row)].value = GOV_POSTAL_CODE.strip(
                                GOV_STATE_REGION_SHORT)
        # Check to see if institution is inactive/closed according to NCES database
        logger.info('Checking to see if Institution is inactive/closed according to NCES database...')
        for cell in ws_uasys['E']:
            if cell.row >= 3:
                institution_govern = str(cell.value)
                cell_prev = int(cell.row) - 1
                try:
                    if cell_prev != 0 and institution_govern.upper() != ws_uasys['E' + str(cell_prev)].value.upper():

                        for look in ws_nces_grab['B']:
                            nces_institution = str(look.value)
                            if nces_institution.upper() == institution_govern.upper():
                                institution_closed = ws_nces_grab['W' + str(
                                    look.row)].value
                                found_two = str(institution_closed).find('-2')
                                if found_two < 0:
                                    ws_uasys['AI' + str(cell.row)
                                             ].value = institution_closed
                except Exception as e:
                    logger.exception(
                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        for cell in ws_uasys['F']:
            if cell.row >= 3:
                governing_address = str(cell.value)
                if governing_address == 'None':
                    one_dapip = str(ws_uasys['R' + str(cell.row)].value)
                    two_oped = str(ws_uasys['S' + str(cell.row)].value)
                    three_iped = str(ws_uasys['T' + str(cell.row)].value)
                    address_one = str(ws_uasys['V' + str(cell.row)].value)
                    address_two = str(ws_uasys['W' + str(cell.row)].value)
                    pobox = str(ws_uasys['X' + str(cell.row)].value)
                    city = str(ws_uasys['Y' + str(cell.row)].value)
                    state = str(ws_uasys['Z' + str(cell.row)].value)
                    zipcode = str(ws_uasys['AB' + str(cell.row)].value)
                    phonenumber = str(ws_uasys['AC' + str(cell.row)].value)

                    ws_uasys['B' + str(cell.row)].value = one_dapip
                    ws_uasys['C' + str(cell.row)].value = two_oped
                    ws_uasys['D' + str(cell.row)].value = three_iped
                    ws_uasys['F' + str(cell.row)].value = address_one
                    ws_uasys['G' + str(cell.row)].value = address_two
                    ws_uasys['H' + str(cell.row)].value = pobox
                    ws_uasys['I' + str(cell.row)].value = city
                    ws_uasys['J' + str(cell.row)].value = state
                    ws_uasys['L' + str(cell.row)].value = zipcode
                    ws_uasys['M' + str(cell.row)].value = phonenumber
        wb_uasys.save(file_location)

    @classmethod
    def clean_governing(cls, wb_uasys, ws_uasys, file_location, full_spellings):
        for cell in ws_uasys['B']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['B' + str(cell.row)].value = "NULL"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['C']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['C' + str(cell.row)].value = "NULL"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['D']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['D' + str(cell.row)].value = "NULL"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['M']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['M' + str(cell.row)].value = "NULL"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        yellow = 'FFFF00'
        red = 'FF6666'
        y_highlight = PatternFill(patternType='solid', fgColor=yellow)
        r_highlight = PatternFill(patternType='solid', fgColor=red)
        for cell in ws_uasys['A']:
            if cell.row >= 3:
                try:
                    org_id = str(cell.value)
                    if org_id != 'AutoGen':
                        ws_uasys['A' + str(cell.row)].value = 'AutoGen'
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['B']:
            if cell.row >= 3:
                try:
                    gov_id = str(cell.value)
                    gov_id_oped = str(ws_uasys['C' + str(cell.row)].value)
                    gov_id_iped = str(ws_uasys['D' + str(cell.row)].value)
                    if not gov_id.isnumeric():
                        ws_uasys['B' + str(cell.row)].fill = r_highlight
                    if not gov_id_oped.isnumeric():
                        ws_uasys['C' + str(cell.row)].fill = r_highlight
                    if not gov_id_iped.isnumeric():
                        ws_uasys['D' + str(cell.row)].fill = r_highlight
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['E']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['E' + str(cell.row)].fill = y_highlight
                    else:
                        ws_uasys['E' + str(cell.row)
                                 ].value = str(cell.value).upper()
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['F']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['F' + str(cell.row)].fill = r_highlight
                    else:
                        ws_uasys['F' + str(cell.row)
                                 ].value = str(cell.value).upper()
                    gov_address = str(
                        ws_uasys['F' + str(cell.row)].value).lower()
                    sep_address = gov_address.split()

                    for key in full_spellings:

                        for index in range(len(sep_address)):
                            word = sep_address[index]
                            if word == key:
                                sep_address[index] = full_spellings[key]
                                gov_address = str(' '.join(sep_address))
                                ws_uasys['F' + str(cell.row)
                                         ].value = gov_address.upper()
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['G']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['G' + str(cell.row)].value = "N/A"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['H']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['H' + str(cell.row)].value = "N/A"
                    elif cell.value == 'None':
                        ws_uasys['H' + str(cell.row)].value = "N/A"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['I']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['I' + str(cell.row)].fill = r_highlight
                    else:
                        ws_uasys['I' + str(cell.row)
                                 ].value = str(cell.value).upper()
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['J']:
            try:
                if cell.row >= 3:
                    region = str(cell.value)
                    if len(region) != 2:
                        ws_uasys['J' + str(cell.row)].fill = r_highlight
            except:
                logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['K']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['K' + str(cell.row)].value = 'USA'
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['L']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['L' + str(cell.row)].fill = r_highlight
                except:
                    logger.exception('Error with cell')
        for cell in ws_uasys['N']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['N' + str(cell.row)].value = "N"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['O']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['O' + str(cell.row)].value = "N/A"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['P']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['P' + str(cell.row)].value = "N/A"
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        wb_uasys.save(file_location)

    @classmethod
    def reconcile_campuslocation(cls, wb_uasys, ws_uasys, file_location, abbrev, ws_data_grab, ws_nces_grab):
        # If CAMPUS_LOCATION_ID is blank then assign the cell AutoGen
        logger.info('If Campus LOCATION_ID is blank then assign the cell(s) AutoGen...')
        for cell in ws_uasys['AK']:
            try:
                if cell.row >= 3:
                    if cell.value is None:
                        ws_uasys['AK' + str(cell.row)].value = "AutoGen"
            except Exception as e:
                logger.exception(
                    f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        # Grabbing the missing cells in CAMP_OFFICIAL_INSTITUTION_NAME from PRIMARY_INSTITUTION_NAME
        logger.info('Grabbing the missing cells in Campus OFFICIAL_INSTITUTION_NAME from PRIMARY_INSTITUTION_NAME...')
        for cell in ws_uasys['AP']:
            try:
                if cell.row >= 3:
                    if cell.value is None:
                        CAMP_OFFICIAL_INSTITUTION_NAME = ws_uasys['U' + str(
                            cell.row)].value
                        ws_uasys['AP' + str(cell.row)
                                 ].value = CAMP_OFFICIAL_INSTITUTION_NAME.upper()
            except AttributeError as e:
                logger.exception(
                    f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        # for column AQ find the additional location for column AP and address for the location
        logger.info('Finding the additional location for column AP and adrress for the location...')
        for cell in ws_uasys['AQ']:
            if cell.row >= 3:
                try:
                    lookup_institution = ws_uasys['AP' + str(cell.row)].value
                    used_cell_ar = False

                    for grab in ws_data_grab['E']:
                        if grab.value.upper == lookup_institution.upper:
                            additional_location = ws_data_grab['D' +
                                                               str(grab.row)].value
                            address_additional_location = ws_data_grab['H' + str(
                                grab.row)].value
                            cell_prev = int(cell.row) - 1
                            prev_additional_location = ws_uasys['AQ' + str(cell_prev)].value if \
                                str(ws_uasys['AQ' + str(cell_prev)].value) != 'N/A' else \
                                ws_uasys['AR' + str(cell_prev)].value
                            used_cell_ar = True if str(
                                ws_uasys['AQ' + str(cell_prev)].value) == 'N/A' else False
                            if additional_location.upper != prev_additional_location.upper \
                                    and additional_location is not None and used_cell_ar is False:
                                ws_uasys['AQ' + str(cell.row)
                                         ].value = str(additional_location).upper
                                ws_uasys['AS' + str(cell.row)].value = str(
                                    address_additional_location).upper
                                wb_uasys.save(file_location)
                            elif used_cell_ar is False:
                                ws_uasys['AQ' + str(cell.row)
                                         ].value = str(lookup_institution).upper
                            elif additional_location.upper != prev_additional_location.upper \
                                    and additional_location is not None and used_cell_ar is True:
                                ws_uasys['AR' + str(cell.row)
                                         ].value = str(additional_location).upper
                                ws_uasys['AS' + str(cell.row)].value = str(
                                    address_additional_location).upper
                                wb_uasys.save(file_location)
                            elif used_cell_ar is True:
                                ws_uasys['AR' + str(cell.row)
                                         ].value = str(lookup_institution).upper
                except Exception as e:
                    logger.exception(
                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        # Get CAMP_CAMPUS_NAME CAMP_OPED_ID and CAMP_IPED_ID from LocationName OpeId and IpedsUnitIds
        logger.info('Getting CAMP_CAMPUS_NAME CAMP_OPED_ID and CAMP_IPED_ID from LocationName OpeId and IpedsUnitIds')
        for cell in ws_uasys['AQ']:
            try:
                if cell.row >= 3:
                    organization_name = str(cell.value) if str(cell.value) != 'N/A' else \
                        str(ws_uasys['AR' + str(cell.row)].value)
                    organization_address = str(
                        ws_uasys['AS' + str(cell.row)].value)
                    logger.debug(f"Populating {organization_name} fields.....")

                    for grab in ws_data_grab['D']:
                        location_name = str(grab.value)
                        location_address = str(
                            ws_data_grab['H' + str(grab.row)].value)
                        string_match = SequenceMatcher(lambda x: x in " \t",
                                                       organization_address, location_address).ratio()
                        if location_name.upper() == organization_name.upper() and string_match >= 0.6:
                            CAMP_DAPID = str(
                                ws_data_grab['A' + str(grab.row)].value)
                            CAMP_OPED_ID = str(
                                ws_data_grab['B' + str(grab.row)].value)
                            CAMP_IPED_ID = str(
                                ws_data_grab['C' + str(grab.row)].value)
                            ws_uasys['AL' + str(cell.row)].value = CAMP_DAPID
                            ws_uasys['AM' + str(cell.row)].value = CAMP_OPED_ID
                            ws_uasys['AN' + str(cell.row)].value = CAMP_IPED_ID
            except Exception as e:
                logger.exception(
                    f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        # for any CAMP_CAMPUS_NAME ids that aren't populated by accred, find data in institution
        logger.info('For any CAMP_CAMPUS_NAME ids that are not populated, find data in Institution...')
        for cell in ws_uasys['AQ']:
            if cell.row >= 3:
                CAMP_CAMPUS_NAME = str(cell.value)
                COIN_dapid = ws_uasys['AL' + str(cell.row)].value
                COIN_opeid = ws_uasys['AM' + str(cell.row)].value
                COIN_ipedid = ws_uasys['AN' + str(cell.row)].value
                if COIN_dapid is None or COIN_opeid is None or COIN_ipedid is None:
                    institution_name = ws_uasys['U' + str(cell.row)].value
                    try:
                        if institution_name == CAMP_CAMPUS_NAME:
                            pop_dapid = str(
                                ws_uasys['R' + str(cell.row)].value)
                            pop_opeid = str(
                                ws_uasys['S' + str(cell.row)].value)
                            pop_ipedid = str(
                                ws_uasys['T' + str(cell.row)].value)
                            ws_uasys['AL' + str(cell.row)].value = pop_dapid
                            ws_uasys['AM' + str(cell.row)].value = pop_opeid
                            ws_uasys['AN' + str(cell.row)].value = pop_ipedid
                    except Exception as e:
                        logger.exception(
                            f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        # If Campus dapid, opeid, ipedid is none then assign the cell NULL
        logger.info('If Campus DAPID, OPEID, IPEDID is none then assign the cell(s) NULL...')
        for cell in ws_uasys['AL']:
            if cell.row >= 3:
                opeid = ws_uasys['AM' + str(cell.row)].value
                ipedid = ws_uasys['AN' + str(cell.row)].value
                try:
                    if cell.value is None:
                        ws_uasys['AL' + str(cell.row)].value = "NULL"
                    if opeid is None:
                        ws_uasys['AM' + str(cell.row)].value = "NULL"
                    if ipedid is None:
                        ws_uasys['AN' + str(cell.row)].value = "NULL"
                except Exception as e:
                    logger.exception(
                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
        # Get CAMP_PO_BOX_LINE and CAMP_PhoneNumberFull from CAMP_CAMPUS_NAME against LocationName fields
        logger.info(
            'Getting Campus PO_BOX_LINE and Campus PhoneNumberFull from Campus CAMPUS_NAME against LocationName fields...')
        for cell in ws_uasys['AQ']:
            if cell.row >= 3:
                organization_name = str(cell.value)

                for grab in ws_data_grab['D']:
                    location_name = str(grab.value)
                    if location_name.upper() == organization_name.upper():
                        CAMP_PhoneNumberFull = str(
                            ws_data_grab['I' + str(grab.row)].value)
                        address_grab = str(
                            ws_data_grab['H' + str(grab.row)].value)
                        address_lst = address_grab.split(', ', maxsplit=7)
                        try:
                            if len(address_lst) == 1:
                                address_grab = address_grab + ", N/A, N/A, N/A, N/A, N/A, N/A"
                            elif len(address_lst) == 2:
                                address_grab = address_grab + ", N/A, N/A, N/A, N/A, N/A"
                            elif len(address_lst) == 3:
                                address_grab = address_grab + ", N/A, N/A, N/A, N/A"
                            elif len(address_lst) == 4:
                                address_grab = address_grab + ", N/A, N/A, N/A"

                            if len(address_grab.split(', ')) == 7:
                                GOV_ADDRESS_LINE_1, temp_LINE_2, temp_POBOX, temp_MUNI, temp_PCODE, temp1_Unknown, \
                                    temp2_Unknown = address_grab.split(', ')

                                if GOV_ADDRESS_LINE_1.startswith('P.O. Box'):
                                    temp_PCODE = temp_POBOX
                                    temp_POBOX = GOV_ADDRESS_LINE_1
                                    GOV_ADDRESS_LINE_1 = str('N/A')

                                if temp_POBOX.startswith('K'):
                                    temp_POBOX = 'N/A'
                                    temp_MUNI = temp_PCODE
                                    temp_PCODE = temp1_Unknown

                                if temp_PCODE.startswith('P.O BOX'):
                                    temp_POBOX = temp_PCODE
                                    temp_PCODE = 'NULL'

                                if not temp_LINE_2.startswith('Suite'):
                                    temp_MUNI = temp_LINE_2
                                    temp_LINE_2 = 'N/A'
                                    temp_PCODE = temp_POBOX
                                    temp_POBOX = 'N/A'

                                CAMP_ADDRESS_LINE_2 = temp_LINE_2.upper()
                                CAMP_PO_BOX_LINE = temp_POBOX.strip('.')
                                CAMP_MUNICIPALITY = temp_MUNI.upper()
                                CAMP_POSTAL_CODE = temp_PCODE.strip(abbrev)

                                ws_uasys['AT' + str(cell.row)
                                         ].value = CAMP_ADDRESS_LINE_2
                                ws_uasys['AU' + str(cell.row)
                                         ].value = CAMP_PO_BOX_LINE
                                ws_uasys['AV' + str(cell.row)
                                         ].value = CAMP_MUNICIPALITY
                                ws_uasys['AY' + str(cell.row)
                                         ].value = CAMP_POSTAL_CODE

                            elif len(address_grab.split(', ')) == 6:
                                GOV_ADDRESS_LINE_1, temp_LINE_2, temp_POBOX, temp_MUNI, temp_PCODE, temp1_Unknown = \
                                    address_grab.split(', ')

                                if GOV_ADDRESS_LINE_1.startswith('P.O. Box'):
                                    temp_PCODE = temp_POBOX
                                    temp_POBOX = GOV_ADDRESS_LINE_1
                                    GOV_ADDRESS_LINE_1 = str('N/A')

                                if temp_POBOX.startswith('K'):
                                    temp_POBOX = 'N/A'
                                    temp_MUNI = temp_PCODE
                                    temp_PCODE = temp1_Unknown

                                if temp_PCODE.startswith('P.O BOX'):
                                    temp_POBOX = temp_PCODE
                                    temp_PCODE = 'NULL'

                                if not temp_LINE_2.startswith('Suite'):
                                    temp_MUNI = temp_LINE_2
                                    temp_LINE_2 = 'N/A'
                                    temp_PCODE = temp_POBOX
                                    temp_POBOX = 'N/A'

                                CAMP_ADDRESS_LINE_2 = temp_LINE_2.upper()
                                CAMP_PO_BOX_LINE = temp_POBOX.strip('.')
                                CAMP_MUNICIPALITY = temp_MUNI.upper()
                                CAMP_POSTAL_CODE = temp_PCODE.strip(abbrev)

                                ws_uasys['AT' + str(cell.row)
                                         ].value = CAMP_ADDRESS_LINE_2
                                ws_uasys['AU' + str(cell.row)
                                         ].value = CAMP_PO_BOX_LINE
                                ws_uasys['AV' + str(cell.row)
                                         ].value = CAMP_MUNICIPALITY
                                ws_uasys['AY' + str(cell.row)
                                         ].value = CAMP_POSTAL_CODE

                            elif len(address_grab.split(', ')) == 5:
                                GOV_ADDRESS_LINE_1, temp_LINE_2, temp_POBOX, temp_MUNI, temp_PCODE = \
                                    address_grab.split(', ')

                                if GOV_ADDRESS_LINE_1.startswith('P.O. Box'):
                                    temp_PCODE = temp_POBOX
                                    temp_POBOX = GOV_ADDRESS_LINE_1
                                    GOV_ADDRESS_LINE_1 = str('N/A')

                                if temp_PCODE.startswith('P.O BOX'):
                                    temp_POBOX = temp_PCODE
                                    temp_PCODE = 'NULL'

                                if not temp_LINE_2.startswith('Suite'):
                                    temp_MUNI = temp_LINE_2
                                    temp_LINE_2 = 'N/A'
                                    temp_PCODE = temp_POBOX
                                    temp_POBOX = 'N/A'

                                CAMP_ADDRESS_LINE_2 = temp_LINE_2.upper()
                                CAMP_PO_BOX_LINE = temp_POBOX.strip('.')
                                CAMP_MUNICIPALITY = temp_MUNI.upper()
                                CAMP_POSTAL_CODE = temp_PCODE.strip(abbrev)

                                ws_uasys['AT' + str(cell.row)
                                         ].value = CAMP_ADDRESS_LINE_2
                                ws_uasys['AU' + str(cell.row)
                                         ].value = CAMP_PO_BOX_LINE
                                ws_uasys['AV' + str(cell.row)
                                         ].value = CAMP_MUNICIPALITY
                                ws_uasys['AY' + str(cell.row)
                                         ].value = CAMP_POSTAL_CODE
                        except Exception as e:
                            logger.exception(
                                f"An exception of type {type(e).__name__} occurred, NULL assigned. Details: {str(e)}")
                            ws_uasys['AT' + str(cell.row)].value = 'NULL'
                            ws_uasys['AU' + str(cell.row)].value = 'NULL'
                            ws_uasys['AV' + str(cell.row)].value = 'NULL'
                            ws_uasys['AY' + str(cell.row)].value = 'NULL'
                        # Phone number is valid regardless of address error
                        ws_uasys['AZ' + str(cell.row)
                                 ].value = CAMP_PhoneNumberFull
        # Grabbing location data from Institution section to campus/location section for main campuses/one location
        logger.info(
            'Grabbing location data from Institution section to campus/location for main campuses/one location...')
        for cell in ws_uasys['AQ']:
            if cell.row >= 3:
                if ws_uasys['AS' + str(cell.row)].value is None:
                    ADDRESS_LINE_1 = ws_uasys['V' + str(cell.row)].value
                    ADDRESS_LINE_2 = ws_uasys['W' + str(cell.row)].value
                    PO_BOX_LINE = ws_uasys['X' + str(cell.row)].value
                    MUNICIPALITY = ws_uasys['Y' + str(cell.row)].value
                    STATE_REGION_SHORT = ws_uasys['Z' + str(cell.row)].value
                    POSTAL_CODE = ws_uasys['AB' + str(cell.row)].value
                    PhoneNumberFull = ws_uasys['AC' + str(cell.row)].value
                    ws_uasys['AQ' + str(cell.row)].value = 'MAIN CAMPUS'
                    ws_uasys['AS' + str(cell.row)].value = ADDRESS_LINE_1
                    ws_uasys['AT' + str(cell.row)].value = ADDRESS_LINE_2
                    ws_uasys['AU' + str(cell.row)].value = PO_BOX_LINE
                    ws_uasys['AV' + str(cell.row)].value = MUNICIPALITY
                    ws_uasys['AW' + str(cell.row)].value = STATE_REGION_SHORT
                    ws_uasys['AX' + str(cell.row)].value = 'USA'
                    ws_uasys['AY' + str(cell.row)].value = POSTAL_CODE
                    ws_uasys['AZ' + str(cell.row)].value = PhoneNumberFull
        # Move/delete substrings from CAMP_ADDRESS_LINE_1 and moving them into respective column row
        logger.info('Moving/Deleteing substrings from Campus ADDRESS_LINE_1...')
        substrings = {
            'Ste',
            'Ste.',
            'STE',
            'STE.',
            'Unit',
            'PO',
            'Po',
            'Suite',
            'suite'
        }
        for cell in ws_uasys['AS']:
            if cell.row >= 3:
                governing_address = str(cell.value).split()

                for index in range(len(governing_address)):
                    word = governing_address[index]

                    for look in substrings:
                        if word == look:
                            GOV_ADDRESS_LINE_2 = str(
                                ' '.join(governing_address[index:len(governing_address)]))
                            found_pobox = GOV_ADDRESS_LINE_2.find('PO Box')
                            if found_pobox == -1:
                                ws_uasys['AU' + str(cell.row)
                                         ].value = GOV_ADDRESS_LINE_2.upper()
                            else:
                                ws_uasys['AT' + str(cell.row)
                                         ].value = GOV_ADDRESS_LINE_2
                                ws_uasys['AU' + str(cell.row)].value = 'N/A'
                            ADDRESS_LINE_1 = str(cell.value)
                            phrase_removal = ADDRESS_LINE_1.find(
                                GOV_ADDRESS_LINE_2)
                            if phrase_removal != -1:
                                ws_uasys['AS' + str(cell.row)].value = ADDRESS_LINE_1.strip(
                                    GOV_ADDRESS_LINE_2)
                        elif word == 'Floor' or word == 'Fl':
                            floor_num = index - 1
                            GOV_ADDRESS_LINE_2 = str(
                                ' '.join(governing_address[floor_num:len(governing_address)]))
                            ws_uasys['AT' + str(cell.row)
                                     ].value = GOV_ADDRESS_LINE_2.upper()
                            ADDRESS_LINE_1 = str(cell.value)
                            phrase_removal = ADDRESS_LINE_1.find(
                                GOV_ADDRESS_LINE_2)
                            if phrase_removal != -1:
                                ws_uasys['AS' + str(cell.row)].value = ADDRESS_LINE_1.strip(
                                    GOV_ADDRESS_LINE_2)
                            break
        # Move/delete substrings from CAMP_ADDRESS_LINE_2
        logger.info('Moving/Deleting substrings from Campus ADDRESS_LINE_2...')
        for cell in ws_uasys['AU']:
            if cell.row >= 3:
                CAMP_PO_BOX_LINE = str(cell.value).split()
                word = CAMP_PO_BOX_LINE[0]
                if word != 'PO' or word != 'N/A':
                    word = word.upper()
                    if word.find('STE') == -1:
                        try:
                            ADDRESS_LINE_2 = ws_uasys['AT' +
                                                      str(cell.row)].value
                            ws_uasys['AU' + str(cell.row)
                                     ].value = ADDRESS_LINE_2
                            ws_uasys['AT' + str(cell.row)].value = 'N/A'
                            ws_uasys['AT2'].value = 'CAMP_ADDRESS_LINE_2'
                        except Exception as e:
                            logger.exception(
                                f"An exception of type {type(e).__name__} occurred, NULL assigned. Details: {str(e)}")
                    else:
                        CAMP_POSTAL_CODE = ws_uasys['AV' + str(cell.row)].value
                        CAMP_MUNICIPALITY = ws_uasys['AU' +
                                                     str(cell.row)].value
                        ws_uasys['AY' + str(cell.row)].value = CAMP_POSTAL_CODE
                        ws_uasys['AV' + str(cell.row)
                                 ].value = CAMP_MUNICIPALITY
                        ws_uasys['AU' + str(cell.row)].value = 'N/A'
        for cell in ws_uasys['AY']:
            if cell.row >= 3:
                POSTAL_CODE = str(cell.value)
                try:
                    if POSTAL_CODE.isalpha() and POSTAL_CODE.upper() != 'N/A':
                        CAMP_MUNICIPALITY = ws_uasys['AY' +
                                                     str(cell.row)].value
                        CAMP_ADDRESS_LINE_2 = ws_uasys['AV' +
                                                       str(cell.row)].value
                        ws_uasys['AT' + str(cell.row)
                                 ].value = CAMP_ADDRESS_LINE_2
                        ws_uasys['AV' + str(cell.row)
                                 ].value = CAMP_MUNICIPALITY
                        ws_uasys['AY' + str(cell.row)].value = ""
                    elif POSTAL_CODE.upper() == 'N/A' or POSTAL_CODE.upper() == 'N/':
                        CAMP_POSTAL_CODE = ws_uasys['AV' + str(cell.row)].value
                        ADDRESS_LINE_2 = ws_uasys['AU' + str(cell.row)].value
                        ws_uasys['AY' + str(cell.row)].value = CAMP_POSTAL_CODE
                        ws_uasys['AT' + str(cell.row)].value = ADDRESS_LINE_2
                        ws_uasys['AU' + str(cell.row)].value = 'N/A'
                        ws_uasys['AU2'].value = 'CAMP_PO_BOX_LINE'
                        ws_uasys['AV' + str(cell.row)].value = ""
                    postal_code_list = str(POSTAL_CODE).split()
                    word = postal_code_list[0]
                    if word.isalpha() and len(word) <= 2:
                        STATE_REGION_SHORT = ''.join(word)
                        ws_uasys['AW' + str(cell.row)
                                 ].value = STATE_REGION_SHORT
                        ws_uasys['AY' + str(cell.row)].value = str(
                            cell.value).strip(STATE_REGION_SHORT)
                except Exception as e:
                    logger.exception(
                        f"An exception of type {type(e).__name__} occurred, NULL assigned. Details: {str(e)}")
        # Check to see if campus is inactive/closed according to NCES database
        logger.info('Check to see if campus is inactive/closed according to NCES database...')
        for cell in ws_uasys['AQ']:
            if cell.row >= 3:
                organization_name = str(cell.value)

                for look in ws_nces_grab['B']:
                    nces_institution = str(look.value)
                    if nces_institution.upper() == organization_name.upper():
                        institution_closed = ws_nces_grab['W' +
                                                          str(look.row)].value
                        found_two = str(institution_closed).find('-2')
                        if found_two < 0:
                            ws_uasys['BC' + str(cell.row)].value = 'Y'
                            ws_uasys['BD' + str(cell.row)
                                     ].value = institution_closed
        # If CAMPUS_RECORD_SOURCE is blank then assign the cell N/A
        logger.info('If CAMPUS_RECORD_SOURCE is blank then assign the cell(s) N/A...')
        for cell in ws_uasys['BE']:
            if cell.row >= 3:
                try:
                    if cell.value is None:
                        ws_uasys['BE' + str(cell.row)].value = "N/A"
                except Exception as e:
                    logger.exception(
                        f"An exception of type {type(e).__name__} occurred, NULL assigned. Details: {str(e)}")
        wb_uasys.save(file_location)

    @classmethod
    def clean_campuslocation(cls, wb_uasys, ws_uasys, file_location, full_spellings):
        for cell in ws_uasys['AK']:
            try:
                if cell.row >= 3:
                    if cell.value is None:
                        ws_uasys['AK' + str(cell.row)].value = 'AutoGen'
            except:
                logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AL']:
            try:
                if cell.row >= 3:
                    if cell.value is None:
                        ws_uasys['AL' + str(cell.row)].value = "NULL"
            except:
                logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AM']:
            try:
                if cell.row >= 3:
                    if cell.value is None:
                        ws_uasys['AM' + str(cell.row)].value = "NULL"
            except:
                logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AN']:
            try:
                if cell.row >= 3:
                    if cell.value is None:
                        ws_uasys['AN' + str(cell.row)].value = "NULL"
            except:
                logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AZ']:
            try:
                if cell.row >= 3:
                    if cell.value is None:
                        ws_uasys['AZ' + str(cell.row)].value = "NULL"
            except:
                logger.exception(f'Error with {cell.coordinate}')
        yellow = 'FFFF00'
        red = 'FF6666'
        y_highlight = PatternFill(patternType='solid', fgColor=yellow)
        r_highlight = PatternFill(patternType='solid', fgColor=red)
        for cell in ws_uasys['AL']:
            try:
                if cell.row >= 3:
                    gov_id = str(cell.value)
                    gov_id_oped = str(ws_uasys['AM' + str(cell.row)].value)
                    gov_id_iped = str(ws_uasys['AN' + str(cell.row)].value)
                    if not gov_id.isnumeric():
                        ws_uasys['AL' + str(cell.row)].fill = r_highlight
                    if not gov_id_oped.isnumeric():
                        ws_uasys['AM' + str(cell.row)].fill = r_highlight
                    if not gov_id_iped.isnumeric():
                        ws_uasys['AN' + str(cell.row)].fill = r_highlight
            except:
                logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AP']:
            try:
                if cell.row >= 3:
                    official_name = str(
                        ws_uasys['U' + str(cell.row)].value).lower()
                    change_name = str(
                        ws_uasys['AP' + str(cell.row)].value).lower()
                    if change_name != official_name:
                        ws_uasys['AP' + str(cell.row)].value = official_name
                    ws_uasys['AP' + str(cell.row)
                             ].value = str(cell.value).upper()
                    if cell.value is None:
                        ws_uasys['AP' + str(cell.row)].fill = r_highlight
            except:
                logger.exception(f'Error with {cell.coordinate}')
        wb_uasys.save(file_location)
        campus_no = (
            "regional",
            "health",
            "center",
            "high",
            "school",
            "technical",
            "inc.",
            "inc",
            "administration",
            "building",
            "office",
            "site",
        )
        for cell in ws_uasys['AQ']:
            try:
                if cell.row >= 3:
                    campus_name = str(
                        ws_uasys['AQ' + str(cell.row)].value).lower()
                    official_name = str(
                        ws_uasys['AP' + str(cell.row)].value).lower()
                    if campus_name == official_name:
                        ws_uasys['AQ' + str(cell.row)].value = "MAIN CAMPUS"
                        ws_uasys['AR' + str(cell.row)].value = "N/A"
                    sep_campus_name = campus_name.split()
                    wb_uasys.save(file_location)

                    for match in campus_no:

                        for index in range(len(sep_campus_name)):
                            word = sep_campus_name[index]
                            if word == match:
                                ws_uasys['AR' + str(cell.row)
                                         ].value = campus_name.upper()
                                ws_uasys['AQ' + str(cell.row)].value = "N/A"
                    wb_uasys.save(file_location)
                    check_na = str(ws_uasys['AQ' + str(cell.row)].value)
                    check_na = check_na.lower()
                    if check_na != 'n/a' or check_na != 'main campus':
                        official_name = re.sub('-', ' - ', official_name)
                        official_name = re.sub(',', ' , ', official_name)
                        sep_official_name = official_name.split()
                        index_list = []

                        for index in range(len(sep_official_name)):
                            remove = sep_official_name[index].lower()
                            word = sep_campus_name[index].lower()
                            if word == remove:
                                index_list.append(index)
                            elif remove == 'at':
                                index_list.append(index)
                                break
                            else:
                                break

                        for index in range(len(sep_campus_name)):
                            remove = sep_campus_name[index]
                            if remove == '-' or remove == ',':
                                index_list.append(index)
                                break
                        remove_element = len(index_list)
                        check_campus = str(sep_campus_name[0])
                        check_official = str(sep_campus_name[0])
                        if remove_element > 0 and check_campus == check_official:

                            i = 1
                            while i <= remove_element:
                                sep_campus_name.pop(0)
                                i += 1
                        if sep_campus_name[0] == '-' or sep_campus_name[0] == ',':
                            sep_campus_name.pop(0)
                        campus = str(' '.join(sep_campus_name)).lower()
                        found_campus = campus.find('campus')
                        if found_campus != -1:
                            ws_uasys['AQ' + str(cell.row)
                                     ].value = campus.upper()
                        else:
                            campus = campus + ' campus'
                            ws_uasys['AQ' + str(cell.row)
                                     ].value = campus.upper()
            except:
                logger.exception(f'Error with {cell.coordinate}')
        wb_uasys.save(file_location)
        for cell in ws_uasys['AR']:
            try:
                if cell.row >= 3:
                    if cell.value is None:
                        ws_uasys['AR' + str(cell.row)].value = "N/A"
                    elif cell.value != "N/A":
                        ws_uasys['AQ' + str(cell.row)].value = "N/A"
            except:
                logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AS']:
            try:
                if cell.row >= 3:
                    ws_uasys['AS' + str(cell.row)
                             ].value = str(cell.value).upper()
                    if cell.value is None:
                        ws_uasys['AS' + str(cell.row)].fill = r_highlight
                    gov_address = str(
                        ws_uasys['AS' + str(cell.row)].value).lower()
                    sep_address = gov_address.split()

                    for key in full_spellings:

                        for index in range(len(sep_address)):
                            word = sep_address[index]
                            if word == key:
                                sep_address[index] = full_spellings[key]
                                gov_address = str(' '.join(sep_address))
                                ws_uasys['AS' + str(cell.row)
                                         ].value = gov_address.upper()
            except:
                logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AT']:
            try:
                if cell.row >= 3:
                    address_one = cell.value
                    if address_one is None:
                        ws_uasys['AT' + str(cell.row)].value = 'N/A'
                    if address_one.find('PO') != -1:
                        ws_uasys['AT' + str(cell.row)].fill = y_highlight
            except:
                logger.exception('Error with cell')
        for cell in ws_uasys['AU']:
            if cell.row >= 3:
                try:
                    address_two = str(cell.value)
                    if cell.value is None:
                        ws_uasys['AU' + str(cell.row)].value = 'N/A'
                        wb_uasys.save(file_location)
                    if address_two.find('PO') == -1 and address_two != 'N/A':
                        ws_uasys['AU' + str(cell.row)].fill = y_highlight
                except:
                    logger.exception('Error with cell')
        for cell in ws_uasys['AV']:
            try:
                if cell.row >= 3:
                    ws_uasys['AV' + str(cell.row)
                             ].value = str(cell.value).upper()
                    if cell.value is None:
                        ws_uasys['AV' + str(cell.row)].fill = r_highlight
            except:
                logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AW']:
            try:
                if cell.row >= 3:
                    region = str(cell.value)
                    if len(region) != 2:
                        ws_uasys['AW' + str(cell.row)].fill = r_highlight
            except:
                logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AX']:
            try:
                if cell.row >= 3:
                    if cell.value is None:
                        ws_uasys['AX' + str(cell.row)].value = 'USA'
            except:
                logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['AY']:
            try:
                if cell.row >= 3:
                    if cell.value is None:
                        ws_uasys['AY' + str(cell.row)].fill = r_highlight
            except:
                logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['BA']:
            if cell.row >= 3:
                try:
                    if cell.value == 'Manually Find' or cell.value is None:
                        ws_uasys['BA' + str(cell.row)].fill = y_highlight
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['BB']:
            if cell.row >= 3:
                try:
                    if cell.value == 'Manually Find' or cell.value is None:
                        ws_uasys['BB' + str(cell.row)].fill = y_highlight
                except:
                    logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['BC']:
            try:
                if cell.row >= 3:
                    if cell.value is None:
                        ws_uasys['BC' + str(cell.row)].value = "N"
            except:
                logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['BD']:
            try:
                if cell.row >= 3:
                    if cell.value is None:
                        ws_uasys['BD' + str(cell.row)].value = "N/A"
            except:
                logger.exception(f'Error with {cell.coordinate}')
        for cell in ws_uasys['BE']:
            try:
                if cell.row >= 3:
                    if cell.value is None:
                        ws_uasys['BE' + str(cell.row)].value = "N/A"
            except:
                logger.exception(f'Error with {cell.coordinate}')
        wb_uasys.save(file_location)

    @classmethod
    def reconcile_nominatim(cls, wb_uasys, ws_uasys, file_location, null_values, gov_field_names, insti_field_names,
                            camp_field_names):
        count: int = int(0)
        for row in ws_uasys.iter_rows(min_row=3, min_col=5, values_only=False):
            count += 1
            cache: list = []
            id_lst: list = []

            # Creating cache of nested lists that will store column letter and n integer
            for cell in row:
                temp: list = []
                cell_content: str = str(cell.value)

                for value in null_values:
                    if cell_content.lower() == value.lower():
                        column: str = str()
                        numbers: str = str()

                        for char in str(cell.coordinate):
                            if char.isalpha():
                                column += char
                            else:
                                numbers += char
                        temp.append(column), temp.append(numbers)
                        cache.append(temp)
            # print(f"This is row {count} cache: " + str(cache))
            # Here is where I will do API requests based on fields that are null_values
            # Skips are what keep track of API call per row for each section of data: reset to false each iteration
            # recursive by row
            skip_gov: bool = bool(False)
            skip_insti: bool = bool(False)
            skip_camp: bool = bool(False)
            run: int = int(0)
            cache_index: int = len(cache) - 1
            while run <= cache_index:
                cache_column = cache[run][0]
                cache_row = cache[run][1]
                try:
                    if not skip_gov:

                        for key in gov_field_names:
                            if cache_column == key:
                                # Assigning variable to call query
                                place_name = str(
                                    ws_uasys['E' + str(cache_row)].value)
                                place_zipcode = str(
                                    ws_uasys['L' + str(cache_row)].value)
                                place_city = str(
                                    ws_uasys['I' + str(cache_row)].value)
                                db_location = str(
                                    ws_uasys['F' + str(cache_row)].value)
                                second_location = str(
                                    ws_uasys['G' + str(cache_row)].value)
                                if second_location != "N/A":
                                    db_location = str(
                                        db_location + ' ' + second_location)
                                place_state = str(
                                    ws_uasys['J' + str(cache_row)].value)

                                try:
                                    missing_data = NominatimIntegration.query_structured(amenity=place_name,
                                                                                         street=db_location,
                                                                                         city=place_city,
                                                                                         state=place_state,
                                                                                         postalcode=place_zipcode)
                                    if missing_data is not None:
                                        ws_uasys['L' + str(cache_row)].value = str(
                                            missing_data['ZipCode'])
                                        ws_uasys['E' + str(cache_row)
                                                 ].value = str(missing_data['Name'])
                                        ws_uasys['I' + str(cache_row)].value = str(
                                            missing_data['Municipality'])
                                        ws_uasys['J' + str(cache_row)
                                                 ].value = str(missing_data['State'])
                                        id_lst.append(
                                            [missing_data['ID'], missing_data['Name']])
                                except Exception as e:
                                    logger.exception(
                                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")

                                skip_gov = True
                                break
                    if not skip_insti:

                        for key in insti_field_names:
                            if cache_column == key:
                                # Assigning variable to call query:
                                place_name = str(
                                    ws_uasys['U' + str(cache_row)].value)
                                place_zipcode = str(
                                    ws_uasys['AB' + str(cache_row)].value)
                                place_city = str(
                                    ws_uasys['Y' + str(cache_row)].value)
                                db_location = str(
                                    ws_uasys['V' + str(cache_row)].value)
                                second_location = str(
                                    ws_uasys['W' + str(cache_row)].value)
                                if second_location != "N/A":
                                    db_location = str(
                                        db_location + ' ' + second_location)
                                place_state = str(
                                    ws_uasys['Z' + str(cache_row)].value)

                                try:
                                    missing_data = NominatimIntegration.query_structured(amenity=place_name,
                                                                                         street=db_location,
                                                                                         city=place_city,
                                                                                         state=place_state,
                                                                                         postalcode=place_zipcode)
                                    if missing_data is not None:
                                        ws_uasys['AB' + str(cache_row)].value = str(
                                            missing_data['ZipCode'])
                                        ws_uasys['U' + str(cache_row)
                                                 ].value = str(missing_data['Name'])
                                        ws_uasys['Y' + str(cache_row)].value = str(
                                            missing_data['Municipality'])
                                        ws_uasys['Z' + str(cache_row)
                                                 ].value = str(missing_data['State'])
                                        id_lst.append(
                                            [missing_data['ID'], missing_data['Name']])
                                except Exception as e:
                                    logger.exception(
                                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")

                                skip_insti = True
                                break
                    if not skip_camp:

                        for key in camp_field_names:
                            if cache_column == key:
                                # Assigning variable to call query:
                                first_name = str(
                                    ws_uasys['AP' + str(cache_row)].value)
                                second_name = str(
                                    ws_uasys['AQ' + str(cache_row)].value)
                                if second_name == 'N/A' or second_name in null_values:
                                    second_name = str(
                                        ws_uasys['AR' + str(cache_row)].value)
                                place_name = str(
                                    first_name + ' ' + second_name)

                                place_zipcode = str(
                                    ws_uasys['AY' + str(cache_row)].value)
                                place_city = str(
                                    ws_uasys['AV' + str(cache_row)].value)
                                db_location = str(
                                    ws_uasys['AS' + str(cache_row)].value)
                                second_location = str(
                                    ws_uasys['AT' + str(cache_row)].value)
                                if second_location != "N/A":
                                    db_location = str(
                                        db_location + ' ' + second_location)
                                place_state = str(
                                    ws_uasys['AW' + str(cache_row)].value)

                                try:
                                    missing_data = NominatimIntegration.query_structured(amenity=place_name,
                                                                                         street=db_location,
                                                                                         city=place_city,
                                                                                         state=place_state,
                                                                                         postalcode=place_zipcode)
                                    if missing_data is not None:
                                        ws_uasys['AY' + str(cache_row)].value = str(
                                            missing_data['ZipCode'])
                                        ws_uasys['AQ' + str(cache_row)
                                                 ].value = str(missing_data['Name'])
                                        ws_uasys['AV' + str(cache_row)].value = str(
                                            missing_data['Municipality'])
                                        ws_uasys['AW' + str(cache_row)
                                                 ].value = str(missing_data['State'])
                                        id_lst.append(
                                            [missing_data['ID'], missing_data['Name']])
                                except Exception as e:
                                    logger.exception(
                                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")

                                skip_camp = True
                                break
                except Exception as e:
                    logger.exception(
                        f"An exception of type {type(e).__name__} occurred. Details: {str(e)}")
                run += 1
            place_id.update_place_ids(id_lst)
        wb_uasys.save(file_location)
